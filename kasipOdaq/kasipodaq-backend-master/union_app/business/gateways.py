import io
import random
import datetime
import sys
import json
import hashlib
import time
import threading

from rest_framework.authtoken.models import Token
from rest_framework.exceptions import NotFound, PermissionDenied

from django.core.exceptions import ObjectDoesNotExist
from django.core.paginator import Paginator
from django.db.models import Q
from django.db import connection
from luhn import *

from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync

from openpyxl import Workbook, load_workbook
from io import BytesIO

from . import entities
from .. import models
from .. import tools


class Gateway:
    def __init__(self, user_role, person_id, context):
        self.user_role = user_role
        self.person_id = person_id
        self.context = context

    def get_book(self, book_class):
        try:
            book = models.Book.objects.get(book_class=book_class)
        except ObjectDoesNotExist:
            raise self.make_exception('book_not_found', {'book_class': book_class})

        return book

    def get_book_record(self, book_class, uid):
        book = self.get_book(book_class)

        try:
            book_record = models.Record.objects.get(book=book, uid=uid)
        except ObjectDoesNotExist:
            raise self.make_exception('book_record_not_found', {'uid': uid})

        return book_record

    def get_book_record_by_id(self, book_class, resource_id):
        book = self.get_book(book_class)

        try:
            book_record = models.Record.objects.get(book=book, pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('record_not_found')

        return book_record

    def get_group_by_key(self, key):
        try:
            group = models.Group.objects.get(key=key)
        except ObjectDoesNotExist:
            raise self.make_exception('record_not_found')

        return group

    def notify(self, person, notification_class, record_id, content, is_personal=False):

        notification = models.Notification()
        notification.attribute_class = notification_class
        notification.record_id = record_id
        notification.content = content

        if is_personal:
            notification.person = person

        else:
            notification.union = models.UnionMember.objects.get(Q(status=1) | Q(status=101), person=person).union

        notification.save()

        group = self.get_group_by_key('notifications')

        try:
            union = models.UnionMember.objects.get(Q(status=1) | Q(status=101), person=person).union
            models.GroupMember.objects.get(group=group, record_id=union.id)

            group_domain = models.GroupDomain()
            group_domain.group = group
            group_domain.domain_name = self.get_book_record('classes', 'applications')
            group_domain.record_id = notification.id

            group_domain.save()
        except ObjectDoesNotExist:
            pass

        return notification.id

    def set_permissions(self, entity, entity_class, is_owner):
        try:
            permission = models.AccessPermission.objects.get(
                role=self.user_role,
                entity_class__uid=entity_class,
                entity_class__book__book_class='classes',
                is_owner=is_owner
            )

            for attribute in entity.__dict__.copy():
                if attribute not in permission.allowed_fields:
                    delattr(entity, attribute)
        except ObjectDoesNotExist:
            pass

    def make_exception(self, exception_key, params=None):
        record_class = self.get_book_record('classes', 'record')
        exception = self.get_book_record('errors', exception_key)

        language_code = self.context['language']

        second_language_code = 'ru' if language_code == 'kk' else 'ru'

        try:
            exception = models.Localization.objects.get(record_id=exception.id,
                                                        attribute_class=record_class,
                                                        key='exception',
                                                        language__code=language_code).value

        except ObjectDoesNotExist:
            try:
                exception = models.Localization.objects.get(record_id=exception.id,
                                                            attribute_class=record_class,
                                                            key='exception',
                                                            language__code=second_language_code).value
            except ObjectDoesNotExist:
                exception = 'key'

        if params is not None:
            for key, value in params:
                exception.replace('{' + key + '}', value)

        return Exception(exception)


class PersonGateway(Gateway):
    def convert(self, person, user=None):
        try:
            setting = person.person_settings.get(person=person)
        except ObjectDoesNotExist:
            setting = models.Setting()
            setting.person = person
            setting.save()

        person_entity = entities.Person()
        person_entity.resource_id = person.id
        person_entity.first_name = person.first_name
        person_entity.family_name = person.family_name
        person_entity.patronymic = person.patronymic
        person_entity.job_position = person.job_position
        person_entity.individual_number = person.individual_number
        person_entity.uid = person.uid

        if user is not None:
            person_entity.phone = user.phone
            person_entity.roles = user.roles

        try:
            person_entity.member_status = person.person_members.get(status__in=[1, 101]).status
        except ObjectDoesNotExist:
            pass

        person_entity.setting = {
            'resource_id': setting.id,
            'enable_notice': setting.enable_notice,
            'language_code': setting.language_code
        }

        person_attributes = person.person_attributes.iterator()

        allowed_fields = {
            'person_birthday': 'birthday',
            'person_sex': 'sex',
            'person_physical_address': 'physical_address',
            'person_email': 'email'
        }

        for person_attr in person_attributes:
            if person_attr.property.uid in allowed_fields:
                setattr(person_entity, allowed_fields[person_attr.property.uid], person_attr.value)

        try:
            person_attribute = person.person_attributes.get(
                property__uid='person_picture',
                property__book__book_class='person_attributes'
            )

            file = models.FileLink.objects.select_related('file').get(
                pk=person_attribute.value,
                file_class__book__book_class='classes',
                file_class__uid='person_attributes',
                type__book__book_class='file_types',
                type__uid='person_picture',
            ).file

            person_entity.picture_uri = f'https://storage.kasipodaq.org/{file.hash}.{file.extension}'
        except ObjectDoesNotExist:
            pass

        try:
            person_attribute = person.person_attributes.get(
                property__uid='person_qr_code',
                property__book__book_class='person_attributes'
            )

            file = models.FileLink.objects.select_related('file').get(
                pk=person_attribute.value,
                file_class__book__book_class='classes',
                file_class__uid='person_attributes',
                type__book__book_class='file_types',
                type__uid='person_qr_code',
            ).file

            person_entity.qr_uri = f'https://storage.kasipodaq.org/{file.hash}.{file.extension}'
        except ObjectDoesNotExist:
            pass

        self.set_permissions(person_entity, 'person', self.person_id == person.id)

        return person_entity

    def delete_picture(self, person_entity):
        try:
            person_attribute = models.PersonAttribute.objects.get(
                person=person_entity.resource_id,
                property__uid='person_picture',
                property__book=self.get_book('person_attributes')
            )

            file_link = models.FileLink.objects.get(
                pk=person_attribute.value
            )

            person_attribute.delete()
            file_link.delete()

        except ObjectDoesNotExist:
            raise self.make_exception('file_not_found')

    def create(self, person_entity, file_gateway):
        person = models.Person()
        person.first_name = person_entity.first_name
        person.family_name = person_entity.family_name
        person.patronymic = person_entity.patronymic
        person.individual_number = person_entity.individual_number
        person.job_position = person_entity.job_position

        person.save()

        settings = models.Setting()
        settings.person = person
        settings.save()

        allowed_fields = {
            'birthday': 'person_birthday',
            'sex': 'person_sex',
            'physical_address': 'person_physical_address',
            'email': 'person_email',
            'picture_id': 'person_picture'
        }

        for key in allowed_fields:
            if hasattr(person_entity, key):
                person_attribute = models.PersonAttribute()
                person_attribute.person = person
                person_attribute.property = self.get_book_record('person_attributes', allowed_fields[key])
                person_attribute.value = getattr(person_entity, key)

                person_attribute.save()

        if hasattr(person_entity, 'picture_id'):
            try:
                file = models.File.objects.get(pk=person_entity.picture_id)
            except ObjectDoesNotExist:
                raise self.make_exception('file_not_found')

            file_link = models.FileLink()
            file_link.file = file
            file_link.type = self.get_book_record('file_types', 'person_picture')
            file_link.file_class = self.get_book_record('classes', 'person_attributes')
            file_link.record_id = person.id

            file_link.save()

        qr_image = tools.QRManager.generate_image(f'{person_entity.first_name} {person_entity.first_name}')
        image_data = tools.QRManager.get_image_content(qr_image)

        file_entity = entities.File()
        file_entity.name = 'qr_code.png'
        file_entity.size = sys.getsizeof(image_data)
        file_entity.content_type = 'image/png'
        file_entity.content = image_data

        file_id = file_gateway.create(file_entity)

        file_link = models.FileLink()
        file_link.file = models.File.objects.get(pk=file_id)
        file_link.type = self.get_book_record('file_types', 'person_qr_code')
        file_link.file_class = self.get_book_record('classes', 'person_attributes')
        file_link.record_id = person.id

        file_link.save()

        person_attr_record = models.Record.objects.get(
            book=self.get_book('person_attributes'),
            uid='person_qr_code'
        )

        person_attribute = models.PersonAttribute()
        person_attribute.person = person
        person_attribute.property = person_attr_record
        person_attribute.value = file_link.id

        person_attribute.save()

        return person.id

    def join_union(self, person_entity, union_entity):
        try:
            person = models.Person.objects.get(pk=person_entity.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('file_not_found')

        try:
            union = models.Union.objects.get(pk=union_entity.resource_id)

            union_member = models.UnionMember()
            union_member.person = person
            union_member.union = union

            if hasattr(union_entity, 'creator'):
                union_member.status = 100 if person_entity.resource_id == union_entity.creator.resource_id else 0

            union_member.save()

            while union.type != 'industry':
                union = union.parent

            person.uid = append('398' + str(union.id).rjust(2, '0') + str(person.id).rjust(10, '0'))
            person.save()

        except ObjectDoesNotExist:
            raise self.make_exception('union_not_found')

        def join_file(file_id, file_type):
            file = models.File.objects.get(pk=file_id)

            file_link = models.FileLink()
            file_link.file = file
            file_link.type = self.get_book_record('file_types', file_type)
            file_link.file_class = self.get_book_record('classes', 'union_members')
            file_link.record_id = union_member.id

            file_link.save()

        if union_member.status == 100:
            join_file(union_entity.application.protocol, 'union_protocol')
            join_file(union_entity.application.position, 'union_position')
            join_file(union_entity.application.statement, 'union_statement')

        try:
            if hasattr(union_entity.application, 'files'):
                for file in union_entity.application.files:
                    file = models.File.objects.get(pk=file.resource_id)

                    file_link = models.FileLink()
                    file_link.file = file
                    file_link.type = self.get_book_record('file_types', 'union_members_docs')
                    file_link.file_class = self.get_book_record('classes', 'union_members')
                    file_link.record_id = union_member.id

                    file_link.save()
        except ObjectDoesNotExist:
            pass

        return union_member.id

    def get_by_id(self, resource_id):
        try:
            person = models.Person.objects.get(pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('person_not_found')

        try:
            user = models.User.objects.get(person=person)
        except ObjectDoesNotExist:
            raise self.make_exception('user_not_found')

        return self.convert(person, user)

    def edit(self, resource_id, person_entity):
        try:
            person = models.Person.objects.get(pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('person_not_found')

        allowed_fields = {
            'birthday': 'person_birthday',
            'sex': 'person_sex',
            'physical_address': 'person_physical_address',
            'email': 'person_email'
        }

        for property_name in person_entity.__dict__.keys():
            if hasattr(person, property_name):
                setattr(person, property_name, getattr(person_entity, property_name))

            if property_name in allowed_fields:
                try:
                    person_attribute = models.PersonAttribute.objects.get(
                        person=person,
                        property__uid=allowed_fields[property_name]
                    )
                    person_attribute.value = getattr(person_entity, property_name)

                except ObjectDoesNotExist:
                    person_attribute = models.PersonAttribute()
                    person_attribute.person = person
                    person_attribute.property = self.get_book_record('person_attributes', allowed_fields[property_name])

                    if property_name == 'birthday':
                        try:
                            person_attribute.value = datetime.datetime.strptime(getattr(person_entity, property_name),
                                                                                '%d-%m-%Y')
                        except ValueError:
                            person_attribute.value = getattr(person_entity, property_name)
                    else:
                        person_attribute.value = getattr(person_entity, property_name)

                person_attribute.save()

        if hasattr(person_entity, 'picture_id'):
            try:
                file = models.File.objects.get(pk=person_entity.picture_id)
            except ObjectDoesNotExist:
                raise self.make_exception('file_not_found')

            file_link = models.FileLink()
            file_link.file = file
            file_link.type = self.get_book_record('file_types', 'person_picture')
            file_link.file_class = self.get_book_record('classes', 'person_attributes')
            file_link.record_id = person.id

            file_link.save()

            try:
                person_attribute = models.PersonAttribute.objects.get(
                    person=person,
                    property__uid='person_picture'
                )

                person_attribute.value = file_link.id
            except ObjectDoesNotExist:
                person_attribute = models.PersonAttribute()
                person_attribute.person = person
                person_attribute.property = self.get_book_record('person_attributes', 'person_picture')
                person_attribute.value = file_link.id

            person_attribute.save()

        person.save()

    def is_union_member(self, person_entity):
        return models.UnionMember.objects.filter(Q(status=1) | Q(status=101) | Q(status=0) | Q(status=100),
                                                 person__id=person_entity.resource_id).exists()


class UserGateway(Gateway):
    def create(self, user_entity):
        try:
            person = models.Person.objects.get(pk=user_entity.person.resource_id)

            user = models.User()
            user.phone = user_entity.phone
            user.password = user_entity.password
            user.person = person
            user.roles = ['user']

            user.save()
        except ObjectDoesNotExist:
            return None

        return user.id

    def get(self, user_entity):
        try:
            login_attempt = models.LoginAttempt.objects.get(ip=user_entity.ip, phone=user_entity.phone,
                                                            updated_date__gte=(
                                                                    datetime.datetime.now() - datetime.timedelta(
                                                                seconds=1)))
            if login_attempt.attempts >= 3:
                raise self.make_exception('many_login_attempts')

        except ObjectDoesNotExist:
            login_attempt = None

        user_exists = models.User.objects.filter(phone=user_entity.phone, status=1).exists()

        if not user_exists:
            raise self.make_exception('user_with_phone_not_found')

        try:
            user = models.User.objects.get(phone=user_entity.phone, password=user_entity.password)
            token, created = Token.objects.get_or_create(user=user)
            models.LoginAttempt.objects.filter(ip=user_entity.ip, phone=user_entity.phone).delete()
        except ObjectDoesNotExist:
            if login_attempt is not None:
                login_attempt.attempts = login_attempt.attempts + 1
                login_attempt.save()
                if login_attempt.attempts >= 3:
                    raise self.make_exception('wrong_password_you_are_blocked')
            else:
                login_attempt = models.LoginAttempt()
                login_attempt.attempts = 1
                login_attempt.ip = user_entity.ip
                login_attempt.phone = user_entity.phone
                login_attempt.save()

            raise self.make_exception('wrong_password')

        auth_user = entities.User()
        auth_user.person = entities.Person()
        auth_user.resource_id = user.id
        auth_user.phone = user_entity.phone
        auth_user.person.resource_id = user.person.id
        auth_user.token = f'Token {token.key}'
        auth_user.roles = user.roles
        auth_user.status = user.status

        return auth_user

    def with_phone_exists(self, phone):
        user_exists = models.User.objects.filter(phone=phone).exists()

        return user_exists

    def restore_password(self, user_entity):
        try:
            user = models.User.objects.get(phone=user_entity.phone)
            user.password = user_entity.password

            user.save()
        except ObjectDoesNotExist:
            return None

    def change_password(self, user_entity):
        old_password = hashlib.sha512(('JYuHliLW6a1yfo1WcC' + user_entity.old_password).encode('utf-8')).hexdigest()

        try:
            user = models.User.objects.get(person=user_entity.person.resource_id, password=old_password)
            user.password = user_entity.password

            user.save()
        except ObjectDoesNotExist:
            raise self.make_exception('the_old_password_is_wrong')


class SettingsGateway(Gateway):
    def __convert(self, settings):
        settings_entity = entities.Settings()
        settings_entity.resource_id = settings.id
        settings_entity.enable_notice = settings.enable_notice
        settings_entity.language_code = settings.language_code

        return settings_entity

    def get(self, person_id):
        try:
            settings = models.Setting.objects.get(person_id=person_id)
        except ObjectDoesNotExist:
            raise self.make_exception('person_not_found')

        return self.__convert(settings)

    def edit(self, settings_entity):
        try:
            settings = models.Setting.objects.get(pk=settings_entity.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('internal_error')

        for key in settings_entity.__dict__.keys():
            if hasattr(settings, key):
                setattr(settings, key, getattr(settings_entity, key))

        settings.save()


class UnionGateway(Gateway):
    def convert(self, union, person_gateway, article_gateway):
        language_code = self.context.get('language', 'ru')
        second_language_code = 'ru' if language_code == 'kk' else 'ru'

        try:
            union_name = models.Localization.objects.get(record_id=union.id,
                                                         language__code=language_code,
                                                         attribute_class__book__book_class='classes',
                                                         attribute_class__uid='union',
                                                         key='name').value
        except ObjectDoesNotExist:
            union_name = models.Localization.objects.get(record_id=union.id,
                                                         language__code=second_language_code,
                                                         attribute_class__book__book_class='classes',
                                                         attribute_class__uid='union',
                                                         key='name').value

        union_entity = entities.Union()
        union_entity.resource_id = union.id
        union_entity.name = union_name
        union_entity.status = union.status
        union_entity.kind = union.type

        full_union_ids = [union.id]
        current_child_ids = list(models.Union.objects.filter(parent_id=union.id).values_list('id', flat=True))

        full_union_ids.extend(current_child_ids)

        while len(current_child_ids) > 0:
            current_child_ids = list(
                models.Union.objects.filter(parent_id__in=current_child_ids).values_list('id', flat=True))
            full_union_ids.extend(current_child_ids)

        member_count = models.UnionMember.objects.filter(union_id__in=full_union_ids, status__in=[1, 101]).count()

        union_entity.member_count = member_count

        count_child = models.Union.objects.filter(parent_id=union.id, status=1).count()
        if count_child == 0:
            union_entity.has_child = False
        else:
            union_entity.has_child = True

        try:
            try:
                city_name = models.Localization.objects.get(record_id=union.id,
                                                            language__code=language_code,
                                                            attribute_class__book__book_class='classes',
                                                            attribute_class__uid='union',
                                                            key='city_name').value
            except ObjectDoesNotExist:
                city_name = models.Localization.objects.get(record_id=union.id,
                                                            language__code=second_language_code,
                                                            attribute_class__book__book_class='classes',
                                                            attribute_class__uid='union',
                                                            key='city_name').value

            union_entity.city_name = city_name
        except ObjectDoesNotExist:
            union_entity.city_name = None

        if union.association is not None:
            try:
                association_name = models.Localization.objects.get(record_id=union.association.id,
                                                                   attribute_class__book__book_class='classes',
                                                                   attribute_class__uid='union',
                                                                   language__code=language_code,
                                                                   key='name').value
            except ObjectDoesNotExist:
                association_name = models.Localization.objects.get(record_id=union.association.id,
                                                                   attribute_class__book__book_class='classes',
                                                                   attribute_class__uid='union',
                                                                   language__code=second_language_code,
                                                                   key='name').value

            union_entity.association_union = entities.Union()
            union_entity.association_union.resource_id = union.association.id if union.parent is not None else None
            union_entity.association_union.name = association_name

        if union.parent is not None:
            try:
                parent_name = models.Localization.objects.get(record_id=union.parent.id,
                                                              attribute_class__book__book_class='classes',
                                                              attribute_class__uid='union',
                                                              language__code=language_code,
                                                              key='name').value
            except ObjectDoesNotExist:
                parent_name = models.Localization.objects.get(record_id=union.parent.id,
                                                              attribute_class__book__book_class='classes',
                                                              attribute_class__uid='union',
                                                              language__code=second_language_code,
                                                              key='name').value

            union_entity.root_union = entities.Union()
            union_entity.root_union.resource_id = union.parent.id if union.parent is not None else None
            union_entity.root_union.name = parent_name

        if union.id > 23:
            try:
                parent_name = models.Localization.objects.get(record_id=union.parent.id,
                                                              attribute_class__book__book_class='classes',
                                                              attribute_class__uid='union',
                                                              language__code=language_code,
                                                              key='name').value
            except ObjectDoesNotExist:
                parent_name = models.Localization.objects.get(record_id=union.parent.id,
                                                              attribute_class__book__book_class='classes',
                                                              attribute_class__uid='union',
                                                              language__code=second_language_code,
                                                              key='name').value

            union_entity.industry = entities.Union()
            union_entity.industry.resource_id = union.parent.id if union.parent is not None else None
            union_entity.industry.name = parent_name

        union_industry = union

        if union.type not in ('industry', 'main_union', 'association', 'tech_support'):
            while union_industry.type != 'industry':
                union_industry = models.Union.objects.get(pk=union_industry.parent.id)

            try:
                parent_name = models.Localization.objects.get(record_id=union_industry.id,
                                                              attribute_class__book__book_class='classes',
                                                              attribute_class__uid='union',
                                                              language__code=language_code,
                                                              key='name').value
            except ObjectDoesNotExist:
                parent_name = models.Localization.objects.get(record_id=union_industry.id,
                                                              attribute_class__book__book_class='classes',
                                                              attribute_class__uid='union',
                                                              language__code=second_language_code,
                                                              key='name').value

            union_entity.industry.resource_id = union_industry.id if union_industry.parent is not None else None
            union_entity.industry.name = parent_name

        file_links = models.FileLink.objects.prefetch_related('file').filter(type__book__book_class='file_types',
                                                                             type__uid='union_members_docs',
                                                                             file_class__book__book_class='classes',
                                                                             file_class__uid='union_members',
                                                                             record_id=union.id)

        file_list = []

        for file_link in file_links:
            file_list.append({
                'name': file_link.file.name,
                'uri': f'https://storage.kasipodaq.org/{file_link.file.hash}.{file_link.file.extension}'
            })

        union_entity.files = file_list

        try:
            file = models.FileLink.objects.select_related('file').get(record_id=union.id,
                                                                      type__book__book_class='file_types',
                                                                      type__uid='union_logo',
                                                                      file_class__book__book_class='classes',
                                                                      file_class__uid='union').file
            union_entity.picture = {
                'name': file.name,
                'uri': f'https://storage.kasipodaq.org/{file.hash}.{file.extension}'
            }
        except ObjectDoesNotExist:
            pass

        try:
            union_member = models.UnionMember.objects.get(Q(status=100) | Q(status=101), union=union)
        except ObjectDoesNotExist:
            union_member = None

        try:
            file_link = models.FileLink.objects.select_related('file').get(record_id=getattr(union_member, 'id', None),
                                                                           type__book__book_class='file_types',
                                                                           type__uid='union_protocol',
                                                                           file_class__book__book_class='classes',
                                                                           file_class__uid='union_members')
            union_entity.protocol = {
                'link_id': file_link.id,
                'name': file_link.file.name,
                'uri': f'https://storage.kasipodaq.org/{file_link.file.hash}.{file_link.file.extension}'
            }
        except ObjectDoesNotExist:
            pass

        try:
            file_link = models.FileLink.objects.select_related('file').get(record_id=getattr(union_member, 'id', None),
                                                                           type__book__book_class='file_types',
                                                                           type__uid='union_position',
                                                                           file_class__book__book_class='classes',
                                                                           file_class__uid='union_members')
            union_entity.position = {
                'link_id': file_link.id,
                'name': file_link.file.name,
                'uri': f'https://storage.kasipodaq.org/{file_link.file.hash}.{file_link.file.extension}'
            }
        except ObjectDoesNotExist:
            pass

        try:
            file_link = models.FileLink.objects.select_related('file').get(record_id=getattr(union_member, 'id', None),
                                                                           type__book__book_class='file_types',
                                                                           type__uid='union_statement',
                                                                           file_class__book__book_class='classes',
                                                                           file_class__uid='union_members')
            union_entity.statement = {
                'link_id': file_link.id,
                'name': file_link.file.name,
                'uri': f'https://storage.kasipodaq.org/{file_link.file.hash}.{file_link.file.extension}'
            }
        except ObjectDoesNotExist:
            pass

        try:
            file_link = models.FileLink.objects.select_related('file').get(record_id=getattr(union_member, 'id', None),
                                                                           type__book__book_class='file_types',
                                                                           type__uid='union_agreement',
                                                                           file_class__book__book_class='classes',
                                                                           file_class__uid='union_members')
            union_entity.agreement = {
                'link_id': file_link.id,
                'name': file_link.file.name,
                'uri': f'https://storage.kasipodaq.org/{file_link.file.hash}.{file_link.file.extension}'
            }
        except ObjectDoesNotExist:
            pass

        try:
            file_link = models.FileLink.objects.select_related('file').get(record_id=union_entity.resource_id,
                                                                           type__book__book_class='file_types',
                                                                           type__uid='union_entry_sample',
                                                                           file_class__book__book_class='classes',
                                                                           file_class__uid='union')
            union_entity.entry_sample = {
                'link_id': file_link.id,
                'name': file_link.file.name,
                'uri': f'https://storage.kasipodaq.org/{file_link.file.hash}.{file_link.file.extension}'
            }
        except ObjectDoesNotExist:
            pass

        try:
            file_link = models.FileLink.objects.select_related('file').get(record_id=union_entity.resource_id,
                                                                           type__book__book_class='file_types',
                                                                           type__uid='union_hold_sample',
                                                                           file_class__book__book_class='classes',
                                                                           file_class__uid='union')
            union_entity.hold_sample = {
                'link_id': file_link.id,
                'name': file_link.file.name,
                'uri': f'https://storage.kasipodaq.org/{file_link.file.hash}.{file_link.file.extension}'
            }
        except ObjectDoesNotExist:
            pass

        try:
            file_link = models.FileLink.objects.select_related('file').get(record_id=union_entity.resource_id,
                                                                           type__book__book_class='file_types',
                                                                           type__uid='union_position_sample',
                                                                           file_class__book__book_class='classes',
                                                                           file_class__uid='union')
            union_entity.position_sample = {
                'link_id': file_link.id,
                'name': file_link.file.name,
                'uri': f'https://storage.kasipodaq.org/{file_link.file.hash}.{file_link.file.extension}'
            }
        except ObjectDoesNotExist:
            pass

        try:
            file_link = models.FileLink.objects.select_related('file').get(record_id=union_entity.resource_id,
                                                                           type__book__book_class='file_types',
                                                                           type__uid='union_protocol_sample',
                                                                           file_class__book__book_class='classes',
                                                                           file_class__uid='union')
            union_entity.protocol_sample = {
                'link_id': file_link.id,
                'name': file_link.file.name,
                'uri': f'https://storage.kasipodaq.org/{file_link.file.hash}.{file_link.file.extension}'
            }
        except ObjectDoesNotExist:
            pass

        try:
            file_link = models.FileLink.objects.select_related('file').get(record_id=union_entity.resource_id,
                                                                           type__book__book_class='file_types',
                                                                           type__uid='union_statement_sample',
                                                                           file_class__book__book_class='classes',
                                                                           file_class__uid='union')
            union_entity.statement_sample = {
                'link_id': file_link.id,
                'name': file_link.file.name,
                'uri': f'https://storage.kasipodaq.org/{file_link.file.hash}.{file_link.file.extension}'
            }
        except ObjectDoesNotExist:
            pass

        file_links = models.FileLink.objects.prefetch_related('file').filter(
            record_id=getattr(union_member, 'id', None),
            type__book__book_class='file_types',
            type__uid='union_sample_application',
            file_class__book__book_class='classes',
            file_class__uid='union_members')

        union_entity.union_sample_applications = []

        for file_link in file_links:
            sample = {
                'resource_id': file_link.file_id,
                'link_id': file_link.id,
                'name': file_link.file.name,
                'uri': f'https://storage.kasipodaq.org/{file_link.file.hash}.{file_link.file.extension}'
            }

            union_entity.union_sample_applications.append(sample)

        allowed_articles = ['about_union', 'about_company']

        for allowed_field in allowed_articles:
            article_entity = entities.Article()
            article_entity.filter_by = {'key': allowed_field, 'union_id': union.id}
            article_entity.person.resource_id = union.creator.id

            articles = article_gateway.get(article_entity, person_gateway)

            if len(articles) > 0:
                setattr(union_entity, allowed_field, articles[0].content)

        bread_crumbs = []

        while union is not None:
            try:
                union_name = models.Localization.objects.get(record_id=union.id,
                                                             language__code=language_code,
                                                             attribute_class__book__book_class='classes',
                                                             attribute_class__uid='union',
                                                             key='name').value
            except ObjectDoesNotExist:
                union_name = models.Localization.objects.get(record_id=union.id,
                                                             language__code=second_language_code,
                                                             attribute_class__book__book_class='classes',
                                                             attribute_class__uid='union',
                                                             key='name').value
            bread_crumbs.append({
                'resource_id': union.id,
                'name': union_name,
                'level': union.type
            })

            union = union.parent

        union_entity.bread_crumbs = bread_crumbs[::-1]

        return union_entity

    def create(self, union_entity):
        try:
            creator = models.Person.objects.get(pk=union_entity.creator.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('person_not_found')

        try:
            root_union = models.Union.objects.get(pk=union_entity.root_union.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('union_not_found')

        if root_union.type == 'industry':
            try:
                root_union = models.Union.objects.get(parent_id=root_union.id,
                                                      association_id=union_entity.association_union.resource_id,
                                                      type='branch')
            except ObjectDoesNotExist:
                pass

        union = models.Union()
        union.creator = creator
        union.parent = root_union

        if hasattr(union_entity, 'association_union') and union_entity.association_union.resource_id is not None:
            try:
                union.association = models.Union.objects.get(pk=union_entity.association_union.resource_id)
            except ObjectDoesNotExist:
                raise self.make_exception('union_not_found')

        union.save()

        for localization in union_entity.localizations:
            language = models.Language.objects.get(code=localization['language_id'])

            union_name = models.Localization()
            union_name.record_id = union.id
            union_name.attribute_class = self.get_book_record('classes', 'union')
            union_name.key = 'name'
            union_name.value = localization['name']
            union_name.language = language

            union_name.save()

        return union.id

    def get_by_id(self, resource_id, person_gateway, article_gateway):
        try:
            union = models.Union.objects.get(pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('union_not_found')

        return self.convert(union, person_gateway, article_gateway)

    def get_list(self, parent_id, union_entity, person_gateway, article_gateway):
        if parent_id is not None:
            root_union = models.Union.objects.get(pk=parent_id)

        else:
            root_union = models.Union.objects.get(parent=None)

        if root_union.type == 'association':
            union_ids_by_parent = list(
                models.Union.objects.filter(association__id=parent_id, type='branch').values_list('id', flat=True))

        else:
            parent_id = parent_id if parent_id is not None else root_union.id

            current_union_ids = [parent_id]

            union_ids_by_parent = []

            while True:
                current_union_ids = list(
                    models.Union.objects.filter(parent__id__in=current_union_ids, status=1).values_list('id',
                                                                                                        flat=True))

                if len(current_union_ids) == 0:
                    break
                else:
                    union_ids_by_parent = union_ids_by_parent + current_union_ids

        if union_entity.search is not None:

            union_ids = list(models.Localization.objects.filter(language__code='ru',
                                                                attribute_class=self.get_book_record('classes',
                                                                                                     'union'),
                                                                key='name',
                                                                value__icontains=union_entity.search).values_list(
                'record_id', flat=True))

            unions = models.Union.objects.filter(id__in=union_ids)

        elif root_union.type == 'association':
            unions = models.Union.objects.filter(pk__in=union_ids_by_parent)

        else:
            unions = models.Union.objects.filter(parent=parent_id)

        unions = unions.filter(status=1)

        unions = unions.filter(**getattr(union_entity, 'filter_by', {})).order_by('sort', '-id')

        paginator = Paginator(unions, union_entity.count)
        unions = paginator.page(union_entity.page_number)

        unions_list = []

        for union in unions:
            unions_list.append(self.convert(union, person_gateway, article_gateway))

        return {
            'unions_list': unions_list,
            'page_number': union_entity.page_number,
            'count': union_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def unions_by_person(self, union_entity, person_gateway, article_gateway, person_id):
        union = models.UnionMember.objects.get(person_id=person_id, status__in=[1, 101]).union

        union_ids = [union.id, union.association_id, 468]

        while union.parent is not None:
            union = union.parent
            union_ids.append(union.id)

        unions = models.Union.objects.filter(pk__in=union_ids).order_by('sort', 'id')

        paginator = Paginator(unions, union_entity.count)
        unions = paginator.page(union_entity.page_number)

        unions_list = []

        for union in unions:
            unions_list.append(self.convert(union, person_gateway, article_gateway))

        return {
            'unions_list': unions_list,
            'page_number': union_entity.page_number,
            'count': union_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def get_industries(self, union_entity, person_gateway, article_gateway):
        industries = models.Union.objects.filter(type='industry').order_by('sort')

        paginator = Paginator(industries, union_entity.count)
        industries = paginator.page(union_entity.page_number)

        industry_list = []

        for industry in industries:
            industry_list.append(self.convert(industry, person_gateway, article_gateway))

        return {
            'unions_list': industry_list,
            'page_number': union_entity.page_number,
            'count': union_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def get_place_associations(self, union_entity, person_gateway, article_gateway):
        root_union = models.Union.objects.get(parent=None)

        place_associations = models.Union.objects.filter(parent=root_union, type='association').order_by('sort', 'id')

        paginator = Paginator(place_associations, union_entity.count)
        place_associations = paginator.page(union_entity.page_number)

        place_association_list = []

        for place_association in place_associations:
            place_association_list.append(self.convert(place_association, person_gateway, article_gateway))

        return {
            'unions_list': place_association_list,
            'page_number': union_entity.page_number,
            'count': union_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def get_union_associations(self, union_entity, person_gateway, article_gateway):
        union_id_list = list(models.Union.objects.filter(type='industry').values_list('id', flat=True))
        union_id_list = list(models.Union.objects.filter(parent__id__in=union_id_list).values_list('id', flat=True))
        branch_id_list = list(models.Union.objects.filter(parent__id__in=union_id_list).values_list('id', flat=True))
        union_id_list = list(models.Union.objects.filter(parent__id__in=branch_id_list).values_list('id', flat=True))

        union_id_list = union_id_list + branch_id_list

        if union_entity.search is not None:
            search_id_list = list(models.Localization.objects.filter(language__code='ru',
                                                                     attribute_class=self.get_book_record('classes',
                                                                                                          'union'),
                                                                     key='name',
                                                                     value__icontains=union_entity.search).values_list(
                'record_id', flat=True))

            union_associations = models.Union.objects.filter(Q(id__in=union_id_list) & Q(id__in=search_id_list))
        else:
            union_associations = models.Union.objects.filter(id__in=union_id_list)

        paginator = Paginator(union_associations, union_entity.count)
        industries = paginator.page(union_entity.page_number)

        industry_list = []

        for industry in industries:
            industry_list.append(self.convert(industry, person_gateway, article_gateway))

        return {
            'unions_list': industry_list,
            'page_number': union_entity.page_number,
            'count': union_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def get_union_application(self, resource_id, person_gateway, article_gateway):
        try:
            union_application = models.UnionMember.objects.get(pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('application_not_found')

        application_entity = entities.UnionApplication()
        application_entity.resource_id = union_application.id
        application_entity.union_name = self.convert(union_application.union, person_gateway, article_gateway).name
        application_entity.person = person_gateway.get_by_id(union_application.person.id)
        application_entity.union = self.convert(union_application.union, person_gateway, article_gateway)
        application_entity.status = union_application.status

        try:
            file_links = models.FileLink.objects.filter(file_class=self.get_book_record('classes', 'union_members'),
                                                        type=self.get_book_record('file_types', 'union_members_docs'),
                                                        record_id=union_application.id)
            file_list = []

            for file_link in file_links:
                file_list.append({
                    'name': file_link.file.name,
                    'uri': f'https://storage.kasipodaq.org/{file_link.file.hash}.{file_link.file.extension}'
                })

            application_entity.files = file_list
        except ObjectDoesNotExist:
            pass

        return application_entity

    def create_union_member(self, person_id, union_master_id):
        union = models.UnionMember.objects.get(person_id=union_master_id, status=101).union
        person = models.Person.objects.get(pk=person_id)

        union_member = models.UnionMember()
        union_member.person = person
        union_member.union = union

        while union.type != 'industry':
            union = union.parent

        person.uid = append('398' + str(union.id).rjust(2, '0') + str(person.id).rjust(10, '0'))
        person.save()

        union_member.save()

        return union_member.id

    def parse_members(self, file_entity):
        members = []
        member_keys = (
            'last_name',
            'first_name',
            'patronymic',
            'birthday',
            'address',
            'phone',
            'email',
            'job_position',
            'sex',
            'uid'
        )

        workbook = load_workbook(filename=BytesIO(file_entity.content))
        worksheet = workbook.worksheets[0]

        rows = list(worksheet.rows)
        data_rows = rows[1:]

        for row in data_rows:
            member = dict()

            for index, key in enumerate(member_keys):
                member[key] = str(row[index].value).strip()

            members.append(member)

        return members

    def parse_children(self, file_entity):
        children = []
        child_keys = (
            'relation',
            'phone',
            'last_name',
            'first_name',
            'birthday',
            'sex',
            'uid'
        )

        workbook = load_workbook(filename=BytesIO(file_entity.content))
        worksheet = workbook.worksheets[0]

        rows = list(worksheet.rows)
        data_rows = rows[1:]

        for row in data_rows:
            child = dict()

            for index, key in enumerate(child_keys):
                child[key] = str(row[index].value).strip()

            children.append(child)

        return children

    def import_members(self, file_entity):
        union = models.UnionMember.objects.get(person_id=self.person_id, status__in=(1, 100, 101)).union

        persons_gateway = PersonGateway(self.person_id, self.user_role, self.context)
        users_gateway = UserGateway(self.person_id, self.user_role, self.context)
        files_gateway = FileGateway(self.person_id, self.user_role, self.context)

        members = self.parse_members(file_entity)

        def create_members(members):
            for member in members:
                if member['phone'] == 'None':
                    continue

                person_entity = entities.Person()
                person_entity.first_name = member['first_name']
                person_entity.family_name = member['last_name']
                person_entity.patronymic = member['patronymic']
                person_entity.birthday = member['birthday'].split(' ')[0]
                person_entity.individual_number = member['uid']
                person_entity.sex = 1 if member['sex'] == 'M' else 0
                person_entity.physical_address = member['address']
                person_entity.job_position = member['job_position']

                user_entity = entities.User()
                user_entity.person = entities.Person()
                user_entity.person.resource_id = persons_gateway.create(person_entity, files_gateway)
                user_entity.phone = '7' + str(member['phone'][1:])
                user_entity.password = '123456'

                try:
                    user = models.User.objects.get(phone=user_entity.phone)
                    user.status = 1
                    user.save()

                    try:
                        union_member = models.UnionMember.objects.get(person=user.person, status__in=(1, 100, 101))
                    except ObjectDoesNotExist:
                        union_member = models.UnionMember()
                        union_member.person = user.person

                    union_member.union = union
                    union_member.status = 101 if union_member.status in (100, 101) else 1

                    union_member.save()
                except ObjectDoesNotExist:
                    users_gateway.create(user_entity)

                    person = models.Person.objects.get(id=user_entity.person.resource_id)

                    union_member = models.UnionMember()
                    union_member.person = person
                    union_member.union = union
                    union_member.join_date = datetime.datetime.now()
                    union_member.status = 1

                    person.uid = append('398' + str(union.id).rjust(2, '0') + str(person.id).rjust(10, '0'))
                    person.save()

                    union_member.save()

        thread = threading.Thread(target=create_members, args=(members,), daemon=True)
        thread.start()

    def import_children(self, file_entity):
        children_gateway = ChildrenGateway(self.person_id, self.user_role, self.context)
        children = self.parse_children(file_entity)

        for child in children:
            child_entity = entities.Child()
            child_entity.family_name = child['last_name']
            child_entity.first_name = child['first_name']
            child_entity.patronymic = None
            child_entity.sex = child['sex'] == 'M'
            child_entity.personal_code = child['uid']
            child_entity.birth_date = child['birthday'].split(' ')[0]

            person = models.Person.objects.get(user__phone=child['phone'])
            child_entity.person = entities.Person()
            child_entity.person.resource_id = person.id

            children_gateway.create(child_entity)

    def is_union_master(self, person_id):
        try:
            models.UnionMember.objects.get(person_id=person_id, status=101)
            return True

        except ObjectDoesNotExist:
            return False

    def set_union_master(self, old_master_id, new_master_id):
        if int(old_master_id) == int(new_master_id):
            return new_master_id

        try:
            union = models.UnionMember.objects.get(person_id=old_master_id, status=101).union
        except ObjectDoesNotExist:
            raise self.make_exception('permission_denied')

        try:
            models.UnionMember.objects.get(person_id=new_master_id, status=1, union_id=union.id)
        except ObjectDoesNotExist:
            raise self.make_exception('internal_error')

        new_master = models.User.objects.get(person_id=new_master_id)
        old_master = models.User.objects.get(person_id=old_master_id)

        new_master.roles = old_master.roles
        new_master.save()

        old_master.roles = '{user}'
        old_master.save()

        old_member = models.UnionMember.objects.get(person_id=old_master_id, status=101)
        old_member.status = 1
        old_member.save()

        new_member = models.UnionMember.objects.get(person_id=new_master_id, status=1)
        new_member.status = 101
        new_member.save()

        return new_master_id

    def get_union_applications(self, person_entity, application_entity, person_gateway, article_gateway, status):
        application_list = []

        union_applications = models.UnionMember.objects.filter(status=status).order_by('-created_date')

        person_id = getattr(person_entity, 'resource_id', None)
        union = models.UnionMember.objects.get(person__id=person_id, status=101).union

        if union.type == 'industry':
            union_list = list(models.Union.objects.filter(parent_id=union.id).values_list('id', flat=True))
            union_list.append(union.id)
        else:
            union_list = [union.id]

        if status == 100:
            union_applications = union_applications.filter(union__parent__id__in=union_list)
        else:
            union_applications = union_applications.filter(union=union)

        paginator = Paginator(union_applications, application_entity.count)
        union_applications = paginator.page(application_entity.page_number)

        for union_application in union_applications:
            application_list.append(self.get_union_application(union_application.id, person_gateway, article_gateway))

        return {
            'unions_list': application_list,
            'page_number': application_entity.page_number,
            'count': application_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def confirm_application(self, resource_id):
        try:
            application = models.UnionMember.objects.get(pk=resource_id)
            application.status = 101 if application.person == application.union.creator else 1
            application.join_date = datetime.datetime.now()

            application.save()
        except ObjectDoesNotExist:
            raise self.make_exception('application_not_found')

        union_name = models.Localization.objects.get(record_id=application.union.id,
                                                     language__code='ru',
                                                     attribute_class=self.get_book_record('classes', 'union'),
                                                     key='name').value

        if application.person == application.union.creator:
            union = application.union
            union.status = 1

            union.save()

            notify_id = self.notify(application.person, self.get_book_record('classes', 'union'), application.union.id,
                                    f'Заявка на создание профсюза "{union_name}" одобрена', True)

            channel_layer = get_channel_layer()

            event = {
                'event': 'EVENT_CONFIRM_APPLICATION',
                'message': 'Заявка на создание профсюза "' + union_name + '" одобрена',
                'record_id': application.union.id,
                'resource_id': notify_id
            }

            async_to_sync(channel_layer.group_send)('person_' + str(application.person.id), {
                'type': 'send_event',
                'event_type': json.dumps(event)
            })

        else:

            notify_id = self.notify(application.person, self.get_book_record('classes', 'union'), application.union.id,
                                    f'Заявка на вступление в профсоюз "{union_name}" одобрена', True)

            channel_layer = get_channel_layer()

            event = {
                'event': 'EVENT_CONFIRM_APPLICATION',
                'message': 'Заявка на создание профсюз "' + union_name + '" одобрена',
                'record_id': application.union.id,
                'resource_id': notify_id
            }

            async_to_sync(channel_layer.group_send)('person_' + str(application.person.id), {
                'type': 'send_event',
                'event_type': json.dumps(event)
            })

        user = models.User.objects.get(person=application.person)

        person = user.person
        person.status = 1
        person.save()

        if application.status == 101:
            user.roles = ['company']

        user.status = 1
        user.save()

    def reject_application(self, resource_id):
        try:
            application = models.UnionMember.objects.get(pk=resource_id)

            union_name = models.Localization.objects.get(record_id=application.union.id,
                                                         language__code='ru',
                                                         attribute_class=self.get_book_record('classes', 'union'),
                                                         key='name').value

            if application.person == application.union.creator:
                notify_id = self.notify(application.person, self.get_book_record('classes', 'union'),
                                        application.union.id,
                                        f'Заявка на создание профсюза "{union_name}" отклонена', True)

                channel_layer = get_channel_layer()

                event = {
                    'event': 'EVENT_CONFIRM_APPLICATION',
                    'message': 'Заявка на создание профсюза "' + union_name + '" отклонена',
                    'record_id': application.union.id,
                    'resource_id': notify_id
                }

                async_to_sync(channel_layer.group_send)('person_' + str(application.person.id), {
                    'type': 'send_event',
                    'event_type': json.dumps(event)
                })

                application.status = 102

            else:
                notify_id = self.notify(application.person, self.get_book_record('classes', 'union'),
                                        application.union.id,
                                        f'Заявка на вступление в профсоюз "{union_name}" отклонена', True)

                channel_layer = get_channel_layer()

                event = {
                    'event': 'EVENT_CONFIRM_APPLICATION',
                    'message': 'Заявка на создание профсюза "' + union_name + '" отклонена',
                    'record_id': application.union.id,
                    'resource_id': notify_id
                }

                async_to_sync(channel_layer.group_send)('person_' + str(application.person.id), {
                    'type': 'send_event',
                    'event_type': json.dumps(event)
                })

                async_to_sync(channel_layer.group_send)('person_' + str(application.person.id), {
                    'type': 'send_event',
                    'event_type': 'EVENT_REJECT_APPLICATION'
                })

                application.status = 2

            application.save()
        except ObjectDoesNotExist:
            raise self.make_exception('application_not_found')

    def delete_application(self, resource_id):
        try:
            application = models.UnionMember.objects.get(pk=resource_id)
            if application.person == application.union.creator:
                application.status = 103
                application.leave_date = datetime.datetime.now()
                application.save()

                models.UnionMember.objects.filter(union_id=application.union_id, status=1).update(
                    leave_date=datetime.datetime.now(), status=3)

            else:
                application.status = 3
                application.leave_date = datetime.datetime.now()
                application.save()
        except ObjectDoesNotExist:
            raise self.make_exception('application_not_found')

    def get_by_member_id(self, resource_id, person_gateway, article_gateway):
        try:
            union_member = models.UnionMember.objects.get(Q(status=1) | Q(status=101), person__id=resource_id)
        except ObjectDoesNotExist:
            return None

        union_entity = self.convert(union_member.union, person_gateway, article_gateway)
        union_entity.join_date = union_member.created_date

        return union_entity

    def get_member_list(self, person_gateway, person_entity, union_id):
        if person_entity.search is not None:

            union_ids = [union_id]

            current_union_ids = [union_id]

            while True:
                current_union_ids = list(
                    models.Union.objects.filter(parent__id__in=current_union_ids, status=1).values_list('id',
                                                                                                        flat=True))

                if len(current_union_ids) == 0:
                    break
                else:
                    union_ids = union_ids + current_union_ids

            person_user_ids = list(
                models.User.objects.filter(phone__icontains=person_entity.search).values_list('person_id', flat=True))

            person_ids = list(models.Person.objects.filter(Q(first_name__icontains=person_entity.search) |
                                                           Q(family_name__icontains=person_entity.search) |
                                                           Q(middle_name__icontains=person_entity.search) |
                                                           Q(patronymic__icontains=person_entity.search) |
                                                           Q(individual_number__icontains=person_entity.search) |
                                                           Q(id__in=person_user_ids)).values_list('id', flat=True))

            member_list = models.UnionMember.objects.filter(Q(status=1) | Q(status=101), union__id__in=union_ids,
                                                            person__id__in=person_ids)

        else:
            member_list = models.UnionMember.objects.filter(Q(status=1) | Q(status=101), union=union_id)

        member_list = member_list.order_by(person_entity.order_by)
        member_list = member_list.order_by('-status')

        paginator = Paginator(member_list, person_entity.count)
        member_list = paginator.page(person_entity.page_number)

        member_entity_list = []

        for member in member_list:
            try:
                user = models.User.objects.get(person=member.person)
                member_entity_list.append(person_gateway.convert(member.person, user))
            except ObjectDoesNotExist:
                pass

        return {
            'members_list': member_entity_list,
            'page_number': person_entity.page_number,
            'count': person_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def get_member_by_id(self, person_gateway, resource_id):
        try:
            member = models.UnionMember.objects.get(Q(status=1) | Q(status=101), person=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('union_member_not_found')

        try:
            user = models.User.objects.get(person=member.person)
        except ObjectDoesNotExist:
            raise self.make_exception('person_not_found')

        member_entity = person_gateway.convert(member.person, user)
        member_entity.status = member.status

        return member_entity

    def edit(self, union_entity, person_gateway, article_gateway):
        try:
            union = models.Union.objects.get(pk=union_entity.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('union_not_found')

        for key in union_entity.__dict__.keys():
            if hasattr(union, key):
                setattr(union, key, getattr(union_entity, key))

        if hasattr(union_entity, 'status') and union_entity.status == 0:
            # Статус 4 - это исключение системой (профсоюз перестал существовать)
            models.UnionMember.objects.filter(union__id=union.id).update(status=4, leave_date=datetime.datetime.now())

        union.save()

        allowed_articles = ['about_union', 'about_company']

        for localization in union_entity.localizations:
            if 'name' in localization:
                language = models.Language.objects.get(code=localization['language_id'])

                try:
                    union_name = models.Localization.objects.get(
                        record_id=union.id,
                        attribute_class=self.get_book_record('classes', 'union'),
                        language=language,
                        key='name'
                    )
                except ObjectDoesNotExist:
                    union_name = models.Localization()
                    union_name.record_id = union.id
                    union_name.language = language
                    union_name.attribute_class = self.get_book_record('classes', 'union')
                    union_name.key = 'name'

                union_name.value = localization['name']
                union_name.save()

            for field in allowed_articles:
                if field in localization:
                    article_entity = entities.Article()
                    article_entity.person.resource_id = union.creator.id
                    article_entity.localizations = [
                        {
                            'language_id': localization['language_id'],
                            'content': localization[field]
                        }
                    ]
                    article_entity.filter_by = {'key': field, 'union_id': union.id}

                    articles = article_gateway.get(article_entity, person_gateway)

                    if len(articles) > 0:
                        article_gateway.edit(articles[0].resource_id, article_entity)
                    else:
                        article_entity.key = field
                        article_entity.union = union_entity

                        article_gateway.create(article_entity)

        def edit_file(file_id, record_id, file_type, file_class, is_list=False):
            if not is_list:
                models.FileLink.objects.filter(record_id=record_id,
                                               file_class=self.get_book_record('classes', file_class),
                                               type=self.get_book_record('file_types', file_type)).delete()

            file_link = models.FileLink()
            file_link.file = models.File.objects.get(pk=file_id)
            file_link.type = self.get_book_record('file_types', file_type)
            file_link.file_class = self.get_book_record('classes', file_class)
            file_link.record_id = record_id
            file_link.save()

        def edit_file_list(file_ids, record_id, file_type, file_class):
            models.FileLink.objects.filter(record_id=record_id,
                                           file_class=self.get_book_record('classes', file_class),
                                           type=self.get_book_record('file_types', file_type)).delete()

            for file_id in file_ids:
                edit_file(file_id, record_id, file_type, file_class, True)

        if hasattr(union_entity, 'picture_id'):
            try:
                edit_file(union_entity.picture_id, union.id, 'union_logo', 'union')
            except ObjectDoesNotExist:
                pass

        if hasattr(union_entity, 'union_protocol_id'):
            try:
                union_member = models.UnionMember.objects.get(status=101, union=union)

                edit_file(union_entity.union_protocol_id, union_member.id, 'union_protocol', 'union_members')
            except ObjectDoesNotExist:
                pass

        if hasattr(union_entity, 'union_position_id'):
            try:
                union_member = models.UnionMember.objects.get(status=101, union=union)

                edit_file(union_entity.union_position_id, union_member.id, 'union_position', 'union_members')
            except ObjectDoesNotExist:
                pass

        if hasattr(union_entity, 'union_statement_id'):
            try:
                union_member = models.UnionMember.objects.get(status=101, union=union)

                edit_file(union_entity.union_statement_id, union_member.id, 'union_statement', 'union_members')
            except ObjectDoesNotExist:
                pass

        if hasattr(union_entity, 'union_agreement_id'):
            try:
                union_member = models.UnionMember.objects.get(status=101, union=union)

                edit_file(union_entity.union_agreement_id, union_member.id, 'union_agreement', 'union_members')
            except ObjectDoesNotExist:
                pass

        if hasattr(union_entity, 'union_sample_applications_list'):
            try:
                union_member = models.UnionMember.objects.get(status=101, union=union)

                edit_file_list(union_entity.union_sample_applications_list, union_member.id, 'union_sample_application',
                               'union_members')
            except ObjectDoesNotExist:
                pass

        if hasattr(union_entity, 'entry_sample_id'):
            try:
                edit_file(union_entity.entry_sample_id, union.id, 'union_entry_sample', 'union')
            except ObjectDoesNotExist:
                pass

        if hasattr(union_entity, 'hold_sample_id'):
            try:
                edit_file(union_entity.hold_sample_id, union.id, 'union_hold_sample', 'union')
            except ObjectDoesNotExist:
                pass

        if hasattr(union_entity, 'position_sample_id'):
            try:
                edit_file(union_entity.position_sample_id, union.id, 'union_position_sample', 'union')
            except ObjectDoesNotExist:
                pass

        if hasattr(union_entity, 'protocol_sample_id'):
            try:
                edit_file(union_entity.protocol_sample_id, union.id, 'union_protocol_sample', 'union')
            except ObjectDoesNotExist:
                pass

        if hasattr(union_entity, 'statement_sample_id'):
            try:
                edit_file(union_entity.statement_sample_id, union.id, 'union_statement_sample', 'union')
            except ObjectDoesNotExist:
                pass

        try:
            if hasattr(union_entity, 'files'):
                models.FileLink.objects.filter(record_id=union.id,
                                               file_class=self.get_book_record('classes', 'union_members'),
                                               type=self.get_book_record('file_types', 'union_members_docs')).delete()

                for file in union_entity.files:
                    file = models.File.objects.get(pk=file.resource_id)

                    file_link = models.FileLink()
                    file_link.file = file
                    file_link.type = self.get_book_record('file_types', 'union_members_docs')
                    file_link.file_class = self.get_book_record('classes', 'union_members')
                    file_link.record_id = union.id

                    file_link.save()
        except ObjectDoesNotExist:
            pass

    def get_by_person_id(self, person_id, person_gateway, article_gateway):
        try:
            union = models.UnionMember.objects.get(Q(status=1) | Q(status=101), person__id=person_id).union
        except ObjectDoesNotExist:
            raise self.make_exception('union_not_found')

        return self.get_by_id(union.id, person_gateway, article_gateway)

    def delete_member(self, resource_id, reason):
        try:
            union_member = models.UnionMember.objects.get(Q(status=1) | Q(status=101), person__id=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('union_member_not_found')

        models.FileLink.objects.filter(record_id=union_member.union.id,
                                       file_class=self.get_book_record('classes', 'union_members'),
                                       type=self.get_book_record('file_types', 'union_members_docs'))

        # Статус 3 - это исключение профсоюзом (сам профсоюз продолжает существовать)
        union_member.status = 3
        union_member.reason = reason
        union_member.save()

        user = models.User.objects.get(person=union_member.person)
        user.roles = ['user']

        user.save()

        union_name = models.Localization.objects.get(record_id=union_member.union.id,
                                                     language__code='ru',
                                                     attribute_class=self.get_book_record('classes', 'union'),
                                                     key='name').value

        notify_id = self.notify(union_member.person, self.get_book_record('classes', 'union'),
                                union_member.union.id,
                                f'Вы были исключены из профсоюза "{union_name}". Причина: {union_member.reason}', True)

        channel_layer = get_channel_layer()

        event = {
            'event': 'EVENT_DELETE_APPLICATION',
            'message': 'Вы были исключены из профсоюза "' + union_name + '". Причина: ' + union_member.reason,
            'record_id': union_member.union.id,
            'resource_id': notify_id
        }

        async_to_sync(channel_layer.group_send)('person_' + str(union_member.person.id), {
            'type': 'send_event',
            'event_type': json.dumps(event)
        })

    def delete_picture(self, person_id, union_id):
        try:
            if union_id is None:
                union_id = models.UnionMember.objects.get(person_id=person_id, status=101).union_id
            else:
                models.UnionMember.objects.get(person_id=person_id, union_id=union_id, status=101)
        except ObjectDoesNotExist:
            raise self.make_exception('permission_denied')

        try:
            file_link = models.FileLink.objects.get(record_id=union_id,
                                                    file_class=self.get_book_record('classes', 'union'),
                                                    type=self.get_book_record('file_types', 'union_logo'))

            file_link.delete()
        except ObjectDoesNotExist:
            raise self.make_exception('file_not_found')


class ShortMessageGateway(Gateway):
    def create(self, sms_entity, ip=None):
        count = models.ShortMessage.objects.filter(phone=sms_entity.phone, created_date__gte=(
                datetime.datetime.now() - datetime.timedelta(minutes=10))).count()

        if count >= 3:
            raise self.make_exception('many_short_message_attempts')

        code = str(random.randint(100000, 999999))

        short_message = models.ShortMessage()
        short_message.phone = sms_entity.phone
        short_message.code = code

        short_message.save()

        # TODO: СДЕЛАТЬ!

        if sms_entity.method == 'create_union_member':
            message = 'Ваш пароль: ' + code

        else:
            message = 'Kasipodaq code: ' + code

        tools.SMSManager.send_sms(short_message.id, sms_entity.phone, message)

        return code

    def confirm(self, phone, code, ip):
        try:
            short_message = models.ShortMessage.objects.get(phone=phone, code=code, status=1, created_date__gte=(
                    datetime.datetime.now() - datetime.timedelta(minutes=5)))
            short_message.status = 0

            short_message.save()
        except ObjectDoesNotExist:
            try:
                message_attempt = models.ShortMessageAttempt.objects.get(phone=phone, ip=ip, created_date__gte=(
                        datetime.datetime.now() - datetime.timedelta(minutes=15)))

                if message_attempt.attempts >= 3:
                    raise self.make_exception('many_short_message_attempts')

                else:
                    message_attempt.attempts += 1
                    message_attempt.save()

            except ObjectDoesNotExist:
                message_attempt = models.ShortMessageAttempt()
                message_attempt.attempts = 1
                message_attempt.ip = ip
                message_attempt.phone = phone
                message_attempt.save()

            raise self.make_exception('code_expired')

        sms_entity = entities.ShortMessage()
        sms_entity.resource_id = short_message.id
        sms_entity.phone = short_message.phone

        return sms_entity

    def has_confirmed(self, phone, sms_code):
        return models.ShortMessage.objects.filter(phone=phone, code=sms_code, status=0).exists()


class NewsGateway(Gateway):
    def create(self, news_entity):
        try:
            person = models.Person.objects.get(pk=news_entity.person.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('person_not_found')

        news = models.News()
        news.person = person
        news.union = models.UnionMember.objects.get(Q(status=1) | Q(status=101), person=person).union

        news.save()

        for localization in news_entity.localizations:
            try:
                language = models.Language.objects.get(code=localization['language_id'])

                news_localization = models.NewsLocalization()
                news_localization.news = news
                news_localization.title = localization['title']
                news_localization.language = language
                news_localization.content = localization['content']
                news_localization.source = localization['source']
                news_localization.short_content = localization['content'] if len(localization['content']) < 100 \
                    else localization['content'][:100]

                news_localization.save()

            except ObjectDoesNotExist:
                pass

        try:
            if hasattr(news_entity, 'picture_id'):
                file = models.File.objects.get(pk=news_entity.picture_id)
                file_link = models.FileLink()
                file_link.file = file
                file_link.type = self.get_book_record('file_types', 'news_picture')
                file_link.file_class = self.get_book_record('classes', 'news')
                file_link.record_id = news.id

                file_link.save()
        except ObjectDoesNotExist:
            pass

        return news.id

    def publish_news(self, resource_id):
        try:
            news = models.News.objects.get(pk=resource_id)
            news_localization = models.NewsLocalization.objects.get(news=news, language__code='ru')
        except ObjectDoesNotExist:
            raise self.make_exception('news_not_found')

        news.published_date = datetime.datetime.now()
        news.status = 1

        news.save()

        notify_id = self.notify(news.person, self.get_book_record('classes', 'news'), news.id,
                                f'Опубликована новость "{news_localization.title}"')

        channel_layer = get_channel_layer()

        event = {
            'event': 'EVENT_NEW_NEWS',
            'message': 'Опубликована новость "' + news_localization.title + '"',
            'record_id': news.id,
            'resource_id': notify_id
        }

        async_to_sync(channel_layer.group_send)('union_' + str(news.union.id), {
            'type': 'send_event',
            'event_type': json.dumps(event)
        })

    def get_list(self, person_entity, person_gateway, news_entity, author_id):
        news_entity_list = []

        person_id = getattr(person_entity, 'resource_id', None)
        union = models.UnionMember.objects.get(Q(status=1) | Q(status=101), person__id=person_id).union

        news_list = models.News.objects.filter(union=union)

        if author_id is not None:
            news_list = news_list.filter(~Q(status=2), person=author_id)
        else:
            news_list = news_list.filter(status=1)

        if news_entity.search is not None:
            news_ids = list(models.NewsLocalization.objects.filter(
                Q(title__icontains=news_entity.search) | Q(short_content__icontains=news_entity.search) | Q(
                    content__icontains=news_entity.search), language__code='ru').values_list('news_id', flat=True))

            news_list = news_list.filter(id__in=news_ids)

        language_code = self.context['language'] if 'language' in self.context else 'ru'

        news_list = news_list.filter(**getattr(news_entity, 'filter_by', {}))
        news_list = news_list.order_by(news_entity.order_by)

        paginator = Paginator(news_list, news_entity.count)
        news_list = paginator.page(news_entity.page_number)

        for news in news_list:
            news_entity_list.append(self.__convert(news, person_gateway, language_code))

        return {
            'news_list': news_entity_list,
            'page_number': news_entity.page_number,
            'count': news_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def get_public_list(self, person_entity, person_gateway, news_entity, author_id):
        news_entity_list = []

        person_id = getattr(person_entity, 'resource_id', None)

        if hasattr(news_entity, 'filter_by') and 'union_id' in news_entity.filter_by:
            union_ids = [news_entity.filter_by['union_id']]
            union = models.Union.objects.get(pk=news_entity.filter_by['union_id'])
            if union.parent is not None and union.type == union.parent.type == 'union':
                union_ids = [union.id, union.parent_id]

            news_list = models.News.objects.filter(union_id__in=union_ids, status=1)

        else:
            try:
                union = models.UnionMember.objects.get(Q(status=1) | Q(status=101), person__id=person_id).union
                if union.parent is not None:
                    union = union.parent
            except ObjectDoesNotExist:
                union = models.Union.objects.get(parent=None)

            news_list = models.News.objects.filter(union=union)

            if hasattr(union, 'association_id') and union.association_id is not None:
                association_union = models.Union.objects.get(pk=union.association_id)
                news_list |= models.News.objects.filter(union=association_union)

            while union.parent is not None:
                union = union.parent

                news_list |= models.News.objects.filter(union=union)

            if author_id is not None:
                news_list = news_list.filter(~Q(status=2), person=author_id)
            else:
                news_list = news_list.filter(status=1)

            news_list = news_list.filter(**getattr(news_entity, 'filter_by', {}))

        if news_entity.search is not None:
            news_ids = list(models.NewsLocalization.objects.filter(
                Q(title__icontains=news_entity.search) | Q(short_content__icontains=news_entity.search) | Q(
                    content__icontains=news_entity.search), language__code='ru').values_list('news_id', flat=True))

            news_list = news_list.filter(id__in=news_ids)

        language_code = self.context['language'] if 'language' in self.context else 'ru'
        news_list = news_list.filter(news_localizations__language__code=language_code)

        news_list = news_list \
            .select_related('person') \
            .prefetch_related('news_localizations')

        news_list = news_list.order_by(news_entity.order_by)

        paginator = Paginator(news_list, news_entity.count)
        news_list = paginator.page(news_entity.page_number)

        for news in news_list:
            news_entity_list.append(self.__convert(news, person_gateway, language_code))

        return {
            'news_list': news_entity_list,
            'page_number': news_entity.page_number,
            'count': news_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def __convert(self, news, person_gateway, language_code='ru'):
        news_entity = entities.News()
        news_entity.resource_id = news.id
        news_entity.person = person_gateway.convert(news.person)

        try:
            news_localization = news.news_localizations.get(language__code=language_code)

            news_entity.title = news_localization.title
            news_entity.content = news_localization.content
            news_entity.source = news_localization.source

        except ObjectDoesNotExist:
            news_entity.title = ''
            news_entity.content = ''
            news_entity.source = ''

        news_entity.is_published = news.status == 1
        news_entity.created_date = news.created_date
        news_entity.updated_date = news.updated_date

        try:
            file = models.FileLink.objects.select_related('file').get(
                file_class__book__book_class='classes',
                file_class__uid='news',
                type__book__book_class='file_types',
                type__uid='news_picture',
                record_id=news.id
            ).file
            news_entity.picture_uri = f'https://storage.kasipodaq.org/{file.hash}.{file.extension}'
        except ObjectDoesNotExist:
            pass

        return news_entity

    def get_by_id(self, resource_id, person_gateway, author_id):
        try:
            news = models.News.objects.get(
                (~Q(status=2) & Q(pk=resource_id) & Q(person=author_id)) | (Q(pk=resource_id) & Q(status=1)))
        except ObjectDoesNotExist:
            raise self.make_exception('news_not_found')

        language_code = self.context['language'] if 'language' in self.context else 'ru'

        news_entity = self.__convert(news, person_gateway, language_code)

        return news_entity

    def delete(self, resource_id):
        try:
            news = models.News.objects.get(pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('news_not_found')

        news.status = 2

        news.save()

    def edit(self, resource_id, news_entity):
        try:
            news = models.News.objects.get(pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('news_not_found')

        for key in news_entity.__dict__.keys():
            if hasattr(news, key):
                setattr(news, key, getattr(news_entity, key))

        news.save()

        for localization in news_entity.localizations:
            try:
                models.Language.objects.get(code=localization['language_id'])

                try:
                    news_localization = models.NewsLocalization.objects.get(news=news,
                                                                            language__code=localization['language_id'])
                except ObjectDoesNotExist:
                    news_localization = models.NewsLocalization()
                    news_localization.news = news

                for key in dict(localization).keys():
                    if hasattr(news_localization, key):
                        if key == 'content':
                            news_localization.content = localization['content']
                            news_localization.short_content = localization['content'] if len(
                                localization['content']) < 100 \
                                else localization['content'][:100]
                        else:
                            setattr(news_localization, key, localization[key])

                news_localization.save()

            except ObjectDoesNotExist:
                pass

        try:
            if hasattr(news_entity, 'picture_id'):
                file = models.File.objects.get(pk=news_entity.picture_id)
                models.FileLink.objects.filter(record_id=news.id,
                                               file_class=self.get_book_record('classes', 'news'),
                                               type=self.get_book_record('file_types', 'news_picture')).delete()

                file_link = models.FileLink()
                file_link.file = file
                file_link.type = self.get_book_record('file_types', 'news_picture')
                file_link.file_class = self.get_book_record('classes', 'news')
                file_link.record_id = news.id

                file_link.save()

        except ObjectDoesNotExist:
            pass


class AppealGateway(Gateway):
    def __convert(self, appeal, person_gateway):
        appeal_entity = entities.Appeal()

        appeal_entity.resource_id = appeal.id
        appeal_entity.type = appeal.type
        appeal_entity.person = person_gateway.convert(appeal.member.person)
        appeal_entity.title = appeal.title
        appeal_entity.content = appeal.content
        appeal_entity.is_published = appeal.status == 0
        appeal_entity.created_date = appeal.created_date
        appeal_entity.updated_date = appeal.updated_date

        try:
            answer = models.Appeal.objects.get(parent=appeal)

            appeal_entity.answer = self.__convert(answer, person_gateway)
        except ObjectDoesNotExist:
            pass

        try:
            file_links = models.FileLink.objects.select_related('file').filter(
                file_class__book__book_class='classes',
                file_class__uid='appeals',
                type__book__book_class='file_types',
                type__uid='appeals_docs',
                record_id=appeal.id
            )

            files_list = []

            for link in file_links:
                files_list.append({
                    'name': link.file.name,
                    'uri': f'https://storage.kasipodaq.org/{link.file.hash}.{link.file.extension}'
                })

            appeal_entity.files = files_list

        except ObjectDoesNotExist:
            pass

        return appeal_entity

    def get_list(self, appeal_entity, person_gateway, person_id, self_appeals):
        appeals = models.Appeal.objects \
            .select_related('member') \
            .select_related('member__person') \
            .filter(~Q(status=2), parent=None)

        if self_appeals:
            appeals = appeals.filter(member__person__id=person_id)

        appeal_entity_list = appeals.filter(**getattr(appeal_entity, 'filter_by', {}))
        appeal_entity_list = appeal_entity_list.order_by(appeal_entity.order_by)

        paginator = Paginator(appeal_entity_list, appeal_entity.count)
        appeal_entity_list = paginator.page(appeal_entity.page_number)

        appeal_list = []

        for appeal in appeal_entity_list:
            appeal_list.append(self.__convert(appeal, person_gateway))

        return {
            'appeal_list': appeal_list,
            'page_number': appeal_entity.page_number,
            'count': appeal_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def get_by_id(self, person_gateway, resource_id):
        try:
            appeal = models.Appeal.objects \
                .select_related('member') \
                .select_related('member__person') \
                .get(parent=None, pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('appeal_not_found')

        appeal_entity = self.__convert(appeal, person_gateway)

        return appeal_entity

    def create(self, appeal_entity, question_id=None):
        try:
            person = models.Person.objects.get(pk=appeal_entity.person.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('person_not_found')

        try:
            union_member = models.UnionMember.objects.get(Q(status=1) | Q(status=101), person=person)
        except ObjectDoesNotExist:
            raise self.make_exception('union_member_not_found')

        appeal = models.Appeal()
        appeal.member = union_member
        appeal.content = appeal_entity.content

        if question_id is not None:
            try:
                question = models.Appeal.objects.get(pk=appeal_entity.question_id, status=0)

                appeal.parent = question
                appeal.type = question.type
                appeal.status = 1

                question.status = 1
                question.save()

                notify_id = self.notify(appeal.parent.member.person, self.get_book_record('classes', 'appeals'),
                                        question.id,
                                        f'На вашу заявку "{question.title}" был дан ответ', True)

                channel_layer = get_channel_layer()

                event = {
                    'event': 'EVENT_APPEAL_ANSWER',
                    'message': 'На вашу заявку "' + question.title + '" был дан ответ',
                    'record_id': question.id,
                    'resource_id': notify_id
                }

                async_to_sync(channel_layer.group_send)('person_' + str(question.member.person.id), {
                    'type': 'send_event',
                    'event_type': json.dumps(event)
                })

            except ObjectDoesNotExist:
                raise self.make_exception('appeal_not_found')

        else:
            try:
                union = models.Union.objects.get(pk=appeal_entity.union.resource_id)
            except ObjectDoesNotExist:
                raise self.make_exception('union_not_found')

            appeal.union = union
            appeal.title = appeal_entity.title
            appeal.type = appeal_entity.type

        appeal.save()

        if hasattr(appeal_entity, 'files'):
            for file in appeal_entity.files:
                file = models.File.objects.get(pk=file.resource_id)

                file_link = models.FileLink()
                file_link.file = file
                file_link.type = self.get_book_record('file_types', 'appeals_docs')
                file_link.file_class = self.get_book_record('classes', 'appeals')
                file_link.record_id = appeal.id

                file_link.save()

        return appeal.id

    def delete(self, resource_id):
        appeals_list = models.Appeal.objects.filter(Q(pk=resource_id) | Q(parent__id=resource_id))
        for appeal in appeals_list:
            appeal.status = 2
            appeal.save()


class ArticleGateway(Gateway):
    def create(self, article_entity):
        try:
            person = models.Person.objects.get(pk=article_entity.person.resource_id)
            user = models.User.objects.get(person=person, status=1)
        except ObjectDoesNotExist:
            raise self.make_exception('person_not_found')

        # if article_entity.key == 'biot' and 'fprk' in user.roles:
        #     raise self.make_exception('permission_denied')

        if article_entity.key in ['about_yntymaq', 'contacts', 'dispute'] and 'fprk' not in user.roles:
            raise self.make_exception('permission_denied')

        article = models.Article()
        article.person = person

        if article_entity.union.resource_id is not None:
            try:
                union = models.Union.objects.get(pk=article_entity.union.resource_id)
                article.union = union

                if union.type == 'main_union':
                    article.sort = 0

                elif union.type == 'association':
                    article.sort = 1

                elif union.type == 'industry':
                    article.sort = 2

                elif union.type == 'branch':
                    article.sort = 3

                elif union.parent.type == 'union':
                    article.sort = 5

                else:
                    article.sort = 4

            except ObjectDoesNotExist:
                raise self.make_exception('union_member_not_found')

        article_type = None

        try:
            article_type = models.Record.objects.get(pk=getattr(article_entity, 'type_id', None))
            article.type = article_type
        except ObjectDoesNotExist:
            pass

        try:
            if article_type is None:
                models.Article.objects.get(key=article_entity.key, union__id=article_entity.union.resource_id)
            else:
                models.Article.objects.get(key=article_entity.key, type=article_type,
                                           union=article_entity.union.resource_id)

            raise self.make_exception('article_exists')
        except ObjectDoesNotExist:
            pass

        article.key = article_entity.key

        article.save()

        for localization in article_entity.localizations:
            try:
                language = models.Language.objects.get(code=localization['language_id'])

                article_localization = models.ArticleLocalization()
                article_localization.article = article
                article_localization.language = language
                article_localization.content = localization['content']

                if 'title' in localization:
                    article_localization.title = localization['title']

                article_localization.save()

            except ObjectDoesNotExist:
                pass

        try:
            if hasattr(article_entity, 'files'):
                for file in article_entity.files:
                    file = models.File.objects.get(pk=file.resource_id)

                    file_link = models.FileLink()
                    file_link.file = file
                    file_link.type = self.get_book_record('file_types', 'article_file')
                    file_link.file_class = self.get_book_record('classes', 'article')
                    file_link.record_id = article.id

                    file_link.save()
        except ObjectDoesNotExist:
            pass

        return article.id

    def edit(self, resource_id, article_entity):
        try:
            article = models.Article.objects.get(pk=resource_id)
            user_creator = models.User.objects.get(person=article.person, status=1)
            current_user = models.User.objects.get(person__id=article_entity.person.resource_id, status=1)
        except ObjectDoesNotExist:
            raise self.make_exception('article_not_found')

        if article.union is not None:
            if article.union.type == 'main_union':
                article.sort = 0

            elif article.union.type == 'association':
                article.sort = 1

            elif article.union.type == 'industry':
                article.sort = 2

            elif article.union.type == 'branch':
                article.sort = 3

            elif article.union.parent.type == 'union':
                article.sort = 5

            else:
                article.sort = 4

        if article.key in ['about_yntymaq', 'contacts', 'dispute'] and 'fprk' not in current_user.roles:
            raise self.make_exception('permission_not_found')

        try:
            person = models.Person.objects.get(pk=article_entity.person.resource_id)
            user = models.User.objects.get(person=person, status=1)
        except ObjectDoesNotExist:
            raise self.make_exception('user_not_found')

        try:
            delattr(article_entity, 'person')
        except:
            pass

        if article.key in ['council', 'biot']:
            if ('fprk' in user.roles and 'fprk' not in user_creator.roles) or (
                    'industry' in user.roles and 'industry' not in user_creator.roles) or (
                    'company' in user.roles and 'company' not in user_creator.roles):
                raise self.make_exception('permission_denied')

        for key in article_entity.__dict__.keys():
            if hasattr(article, key):
                setattr(article, key, getattr(article_entity, key))

        article.save()

        for localization in article_entity.localizations:
            try:
                models.Language.objects.get(code=localization['language_id'])

                try:
                    article_localization = models.ArticleLocalization.objects.get(article=article,
                                                                                  language__code=localization[
                                                                                      'language_id'])
                except ObjectDoesNotExist:
                    article_localization = models.ArticleLocalization()
                    article_localization.article = article

                for key in dict(localization).keys():
                    if hasattr(article_localization, key):
                        setattr(article_localization, key, localization[key])

                article_localization.save()

            except ObjectDoesNotExist:
                pass

        try:
            if hasattr(article_entity, 'files'):
                models.FileLink.objects.filter(file_class=self.get_book_record('classes', 'article'),
                                               type=self.get_book_record('file_types', 'article_file'),
                                               record_id=article.id).delete()

                for file in article_entity.files:
                    file = models.File.objects.get(pk=file.resource_id)

                    file_link = models.FileLink()
                    file_link.file = file
                    file_link.type = self.get_book_record('file_types', 'article_file')
                    file_link.file_class = self.get_book_record('classes', 'article')
                    file_link.record_id = article.id

                    file_link.save()
        except ObjectDoesNotExist:
            pass

    def __convert(self, article, person_gateway, language_code='ru'):
        second_language_code = 'ru' if language_code == 'kk' else 'ru'

        try:
            article_localization = article.article_localization.get(language__code=language_code)

        except ObjectDoesNotExist:
            article_localization = article.article_localization.get(language__code=second_language_code)

        article_entity = entities.Article()
        article_entity.resource_id = article.id
        article_entity.key = article.key
        # article_entity.person = person_gateway.convert(article.person)
        article_entity.title = article_localization.title
        article_entity.content = article_localization.content
        article_entity.created_date = article.created_date
        article_entity.updated_date = article.updated_date
        article_entity.status = article.status

        if article.union is not None:
            union_entity = entities.Union()
            union_entity.resource_id = article.union.id

            try:
                union_entity.name = models.Localization.objects.get(record_id=article.union.id,
                                                                    language__code=language_code,
                                                                    attribute_class__book__book_class='classes',
                                                                    attribute_class__uid='union',
                                                                    key='name').value

            except ObjectDoesNotExist:
                union_entity.name = models.Localization.objects.get(record_id=article.union.id,
                                                                    language__code=second_language_code,
                                                                    attribute_class__book__book_class='classes',
                                                                    attribute_class__uid='union',
                                                                    key='name').value

            article_entity.union = union_entity

        try:
            file_links = models.FileLink.objects.prefetch_related('file').filter(file_class__book__book_class='classes',
                                                                                 file_class__uid='article',
                                                                                 type__book__book_class='file_types',
                                                                                 type__uid='article_file',
                                                                                 record_id=article.id)

            files_list = []

            for link in file_links:
                files_list.append({
                    'name': link.file.name,
                    'uri': f'https://storage.kasipodaq.org/{link.file.hash}.{link.file.extension}'
                })

            article_entity.files = files_list

        except ObjectDoesNotExist:
            pass

        return article_entity

    def get(self, article_entity, person_gateway):
        if hasattr(article_entity, 'self') and article_entity.self:
            try:
                union_member = models.UnionMember.objects.get(Q(status=1) | Q(status=101),
                                                              person__id=article_entity.person.resource_id)
            except ObjectDoesNotExist:
                raise self.make_exception('union_member_not_found')

            article_entity.filter_by['union_id'] = union_member.union.id

        if hasattr(article_entity,
                   'parent_articles') and article_entity.parent_articles and 'union_id' in article_entity.filter_by:
            union = models.Union.objects.get(pk=article_entity.filter_by['union_id'])
            article_entity_list = models.Article.objects.filter(union=union)

            while union.parent is not None:
                union = union.parent

                article_entity_list |= models.Article.objects.filter(union=union)

            del article_entity.filter_by['union_id']

            article_entity_list = article_entity_list.filter(**getattr(article_entity, 'filter_by', {}))

        else:
            article_entity_list = models.Article.objects.filter(**getattr(article_entity, 'filter_by', {}))

        article_entity_list = article_entity_list.order_by('sort', 'union_id')

        article_entity_list.prefetch_related('article_localization').select_related('union', 'person')

        article_list = []

        language_code = self.context['language'] if 'language' in self.context else 'ru'

        for article in article_entity_list:
            article_list.append(self.__convert(article, person_gateway, language_code))

        return article_list


class LegislationGateway(Gateway):
    def create(self, legislation_entity):

        legislation = models.Legislation()

        if legislation_entity.parent_id is not None:
            try:
                legislation.parent = models.Legislation.objects.get(pk=legislation_entity.parent_id)
            except ObjectDoesNotExist:
                raise self.make_exception('legislation_not_found')

        legislation.save()

        for localization in legislation_entity.localizations:
            try:
                language = models.Language.objects.get(code=localization['language_id'])

                legislation_localization = models.LegislationLocalization()
                legislation_localization.legislation = legislation
                legislation_localization.title = localization['title']
                legislation_localization.language = language

                if 'content' in localization:
                    legislation_localization.content = localization['content']

                legislation_localization.save()

            except ObjectDoesNotExist:
                pass

        return legislation.id

    def edit(self, resource_id, legislation_entity):
        try:
            legislation = models.Legislation.objects.get(pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('legislation_not_found')

        for localization in legislation_entity.localizations:
            try:
                language = models.Language.objects.get(code=localization['language_id'])
                try:
                    legislation_localization = models.LegislationLocalization.objects.get(legislation=legislation,
                                                                                          language__code=localization[
                                                                                              'language_id'])

                except ObjectDoesNotExist:
                    legislation_localization = models.LegislationLocalization()

                legislation_localization.legislation = legislation
                legislation_localization.title = localization['title']
                legislation_localization.language = language

                if 'content' in localization:
                    legislation_localization.content = localization['content']

                legislation_localization.save()

            except ObjectDoesNotExist:
                pass

    def get_list(self, legislation_entity):
        if 'parent_id' in legislation_entity.filter_by:
            legislation_entity_list = models.Legislation.objects.filter(
                **getattr(legislation_entity, 'filter_by', {}))

        else:
            legislation_entity_list = models.Legislation.objects.filter(parent=None)

        if legislation_entity.search is not None:
            legislation_ids = list(models.LegislationLocalization.objects.filter(
                Q(content__icontains=legislation_entity.search) | Q(title__icontains=legislation_entity.search),
                language__code='ru').values_list(
                'legislation_id', flat=True))

            legislation_entity_list = models.Legislation.objects.filter(id__in=legislation_ids)

        legislation_entity_list = legislation_entity_list.prefetch_related('legislation_localization').filter(
            **getattr(legislation_entity, 'filter_by', {}))
        legislation_entity_list = legislation_entity_list.order_by(legislation_entity.order_by)

        paginator = Paginator(legislation_entity_list, legislation_entity.count)
        legislation_entity_list = paginator.page(legislation_entity.page_number)

        legislation_list = []

        language_code = self.context['language'] if 'language' in self.context else 'ru'

        for legislation in legislation_entity_list:
            legislation_list.append(self.__convert(legislation, language_code))

        return {
            'legislation_list': legislation_list,
            'page_number': legislation_entity.page_number,
            'count': legislation_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def get_by_id(self, resource_id):
        try:
            legislation = models.Legislation.objects.prefetch_related('legislation_localization').get(pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('legislation_not_found')

        language_code = self.context['language'] if 'language' in self.context else 'ru'

        legislation_entity = self.__convert(legislation, language_code)

        return legislation_entity

    def __convert(self, legislation, language_code='ru'):
        second_language_code = 'ru' if language_code == 'kk' else 'ru'
        try:
            legislation_localization = legislation.legislation_localization.get(language__code=language_code)

        except ObjectDoesNotExist:
            legislation_localization = legislation.legislation_localization.get(language__code=second_language_code)

        legislation_entity = entities.Legislation()
        legislation_entity.resource_id = legislation.id

        if legislation.parent is not None:
            try:
                parent_localization = legislation.parent.legislation_localization.get(language__code=language_code)

            except ObjectDoesNotExist:
                parent_localization = legislation.parent.legislation_localization.get(
                    language__code=second_language_code)

            legislation_entity.parent = entities.Legislation()
            legislation_entity.parent.title = parent_localization.title
            legislation_entity.parent.created_date = legislation.parent.created_date
            legislation_entity.parent.updated_date = legislation.parent.updated_date

        legislation_entity.title = legislation_localization.title
        legislation_entity.content = legislation_localization.content
        legislation_entity.created_date = legislation.created_date
        legislation_entity.updated_date = legislation.updated_date

        legislation_children_list = []

        legislation_child_list = models.Legislation.objects.filter(parent=legislation)

        for legislation_child in legislation_child_list:
            try:
                legislation_localization = legislation_child.legislation_localization.get(language__code=language_code)
            except ObjectDoesNotExist:
                legislation_localization = legislation_child.legislation_localization.get(
                    language__code=second_language_code)

            legislation_child_entity = entities.Legislation()
            legislation_child_entity.resource_id = legislation_child.id
            legislation_child_entity.title = legislation_localization.title
            legislation_child_entity.created_date = legislation_child.created_date
            legislation_child_entity.updated_date = legislation_child.updated_date

            legislation_children_list.append(legislation_child_entity)

        legislation_entity.children = legislation_children_list

        bread_crumbs = []

        while legislation is not None:
            try:
                legislation_localization = models.LegislationLocalization.objects.get(legislation=legislation,
                                                                                      language__code=language_code)
            except ObjectDoesNotExist:
                legislation_localization = models.LegislationLocalization.objects.get(legislation=legislation,
                                                                                      language__code=second_language_code)
            bread_crumbs.append({
                'resource_id': legislation.id,
                'name': legislation_localization.title,
            })

            legislation = legislation.parent

        legislation_entity.bread_crumbs = bread_crumbs[::-1]

        return legislation_entity

    def delete(self, resource_id):
        try:
            models.Legislation.objects.get(pk=resource_id).delete()
        except ObjectDoesNotExist:
            raise self.make_exception('legislation_not_found')


class FileGateway(Gateway):
    def convert(self, file):
        file_entity = entities.File()
        file_entity.filename = file.name
        file_entity.extension = file.extension
        file_entity.size = file.size
        file_entity.content_type = file.content_type
        file_entity.uri = f'https://storage.kasipodaq.org/{file.hash}.{file.extension}'

        return file_entity

    def get_by_hash(self, file_hash, person_entity):
        try:
            file = models.File.objects.filter(hash=file_hash).first()
        except ObjectDoesNotExist:
            raise NotFound('Такого файла не существует')

        # if file.author is not None:
        #     if person_entity is not None:
        #         is_owner = person_entity.resource_id == file.author.id
        #     else:
        #         is_owner = False
        #
        #     try:
        #         permission = models.AccessPermission.objects.get(
        #             role=self.user_role,
        #             entity_class__uid='file',
        #             entity_class__directory__directory_class='entity_classes',
        #             is_owner=is_owner
        #         )
        #
        #         if not permission.has_access:
        #             raise PermissionDenied
        #     except ObjectDoesNotExist:
        #         pass

        return self.convert(file)

    def create(self, file_entity):
        file = models.File()
        file.hash = file_entity.hash
        file.extension = file_entity.extension
        file.name = file_entity.name
        file.size = file_entity.size

        file.save()

        ftp_manager = tools.FTPManager()
        ftp_manager.upload_file('public', f'{file.hash}.{file_entity.extension}', file_entity.content)

        return file.id

    def delete(self, record_id, file_class_id):
        try:
            models.FileLink.objects.get(file_class_id=file_class_id, record_id=record_id).delete()
        except ObjectDoesNotExist:
            raise self.make_exception('file_not_found')

    def delete_link_by_id(self, link_id):
        try:
            models.FileLink.objects.get(id=link_id).delete()
        except ObjectDoesNotExist:
            pass


class DisputeGateway(Gateway):
    def __convert(self, dispute, language_code='ru'):
        dispute_entity = entities.Dispute()
        dispute_entity.resource_id = dispute.id
        dispute_entity.start_date = dispute.start_date
        dispute_entity.finish_date = dispute.finish_date
        dispute_entity.resolved = dispute.status == 1

        second_language_code = 'ru' if language_code == 'kk' else 'ru'

        try:
            dispute_localization = models.DisputeLocalization.objects.get(dispute_id=dispute.id,
                                                                          language_id=language_code)
        except:
            dispute_localization = models.DisputeLocalization.objects.get(dispute_id=dispute.id,
                                                                          language_id=second_language_code)

        dispute_entity.title = dispute_localization.title
        dispute_entity.thesis = dispute_localization.thesis
        dispute_entity.solution = dispute_localization.solution

        return dispute_entity

    def get_list(self, dispute_entity, person_id, self_disputes):
        try:
            union_member = models.UnionMember.objects.get(Q(status=1) | Q(status=101), person__id=person_id)
        except ObjectDoesNotExist:
            union_member = None

        if self_disputes:
            union = union_member.union
            disputes = models.Dispute.objects.filter(~Q(status=2), member__union=union)

            if hasattr(union, 'association_id') and union.association_id is not None:
                association_union = models.Union.objects.get(pk=union.association_id)
                disputes |= models.Dispute.objects.filter(member__union=association_union)

            while union.parent is not None:
                union = union.parent

                disputes |= models.Dispute.objects.filter(~Q(status=2), member__union=union)

        elif dispute_entity.union_id is not None:
            disputes = models.Dispute.objects.filter(~Q(status=2), member__union__id=dispute_entity.union_id)

        else:
            disputes = models.Dispute.objects.filter(~Q(status=2))

        dispute_entity_list = disputes.filter(**getattr(dispute_entity, 'filter_by', {}))
        dispute_entity_list = dispute_entity_list.order_by(dispute_entity.order_by)

        paginator = Paginator(dispute_entity_list, dispute_entity.count)
        dispute_entity_list = paginator.page(dispute_entity.page_number)

        dispute_list = []

        language_code = self.context['language'] if 'language' in self.context else 'ru'

        for dispute in dispute_entity_list:
            dispute_list.append(self.__convert(dispute, language_code))

        return {
            'dispute_list': dispute_list,
            'page_number': dispute_entity.page_number,
            'count': dispute_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def get_by_id(self, resource_id):
        try:
            dispute = models.Dispute.objects.get(~Q(status=2), pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('dispute_not_found')

        language_code = self.context['language'] if 'language' in self.context else 'ru'

        return self.__convert(dispute, language_code)

    def create(self, dispute_entity):
        try:
            member = models.UnionMember.objects.get(Q(status=1) | Q(status=101),
                                                    person__id=dispute_entity.person.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('union_member_not_found')

        dispute = models.Dispute()
        dispute.member = member
        dispute.category = models.Record.objects.get(pk=dispute_entity.category_id)

        for key in dispute_entity.__dict__.keys():
            if hasattr(dispute, key):
                setattr(dispute, key, getattr(dispute_entity, key))

        dispute.save()

        for localization in dispute_entity.localizations:
            dispute_localization = models.DisputeLocalization()
            dispute_localization.dispute = dispute
            dispute_localization.language_id = localization['language_id']
            dispute_localization.title = localization['title']
            dispute_localization.thesis = localization['thesis']
            dispute_localization.solution = localization['solution']

            dispute_localization.save()

        # TODO: Сделать
        # notify_id = self.notify(member.person, self.get_book_record('classes', 'disputes'), dispute.id,
        #                         f'Опубликован новый спор "{dispute_entity.title}"')
        #
        # channel_layer = get_channel_layer()
        #
        # event = {
        #     'event': 'EVENT_NEW_DISPUTE',
        #     'message': 'Опубликован новый спор "' + dispute_entity.title + '"',
        #     'record_id': dispute.id,
        #     'resource_id': notify_id
        # }
        #
        # async_to_sync(channel_layer.group_send)('union_' + str(dispute.member.union.id), {
        #     'type': 'send_event',
        #     'event_type': json.dumps(event)
        # })

        return dispute.id

    def edit(self, dispute_entity):
        try:
            dispute = models.Dispute.objects.get(~Q(status=2), pk=dispute_entity.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('dispute_not_found')

        for localization in dispute_entity.localizations:
            try:
                dispute_localization = models.DisputeLocalization.objects.get(language_id=localization['language_id'],
                                                                              dispute_id=dispute.id)
            except ObjectDoesNotExist:
                dispute_localization = models.DisputeLocalization()
                dispute_localization.dispute = dispute
                dispute_localization.language_id = localization['language_id']

            dispute_localization.title = localization['title']
            dispute_localization.thesis = localization['thesis']

            if 'solution' in localization:
                dispute_localization.solution = localization['solution']

            dispute_localization.save()

    def resolve_dispute(self, dispute_entity):
        try:
            dispute = models.Dispute.objects.get(~Q(status=2), pk=dispute_entity.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('dispute_not_found')

        dispute.finish_date = datetime.datetime.now()
        dispute.status = 1

        dispute.save()

    def delete(self, resource_id):
        try:
            dispute = models.Dispute.objects.get(~Q(status=2), pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('dispute_not_found')

        dispute.status = 2
        dispute.save()


class ChildrenGateway(Gateway):
    def __convert(self, child):

        child_entity = entities.Child()
        child_entity.resource_id = child.id
        child_entity.family_name = child.family_name
        child_entity.first_name = child.first_name
        child_entity.patronymic = child.patronymic
        child_entity.sex = child.sex
        child_entity.personal_code = child.personal_code
        child_entity.birth_date = child.birth_date
        child_entity.age = int((datetime.date.today() - child.birth_date.date()).days / 365.2425)

        return child_entity

    def get_list(self, child_entity):
        children = models.Children.objects.filter(person__id=child_entity.person.resource_id)

        children = children.filter(**getattr(child_entity, 'filter_by', {}))
        children = children.order_by(child_entity.order_by)

        paginator = Paginator(children, child_entity.count)
        children = paginator.page(child_entity.page_number)

        child_list = []

        for child in children:
            child_list.append(self.__convert(child))

        return {
            'child_list': child_list,
            'page_number': child_entity.page_number,
            'count': child_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def get_by_id(self, child_entity):
        try:
            child = models.Children.objects.get(pk=child_entity.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('child_not_found')

        return self.__convert(child)

    def create(self, child_entity):
        try:
            person = models.Person.objects.get(pk=child_entity.person.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('union_member_not_found')

        child = models.Children()
        child.person = person
        child.family_name = child_entity.family_name
        child.first_name = child_entity.first_name
        child.patronymic = child_entity.patronymic
        child.sex = child_entity.sex
        child.personal_code = child_entity.personal_code
        child.birth_date = child_entity.birth_date

        child.save()

        return child.id

    def edit(self, child_entity):
        try:
            child = models.Children.objects.get(pk=child_entity.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('child_not_found')

        try:
            master_union = models.UnionMember.objects.get(person_id=child_entity.person.resource_id,
                                                          status__in=[1, 101]).union

            child_union = models.UnionMember.objects.get(person_id=child.person_id, status__in=[1, 101]).union

            if master_union.id != child_union.id and child_entity.person.resource_id != child.person_id:
                raise self.make_exception('permission_not_found')

        except ObjectDoesNotExist:
            if child_entity.person.resource_id != child.person_id:
                raise self.make_exception('permission_not_found')

        for key in child_entity.__dict__.keys():
            if hasattr(child, key):
                setattr(child, key, getattr(child_entity, key))

        child.save()

    def delete(self, resource_id):
        models.Children.objects.filter(pk=resource_id).delete()


class PartnerGateway(Gateway):
    def convert(self, partner, language_code='ru'):
        partner_entity = entities.Partner()
        partner_entity.resource_id = partner.id

        second_language_code = 'ru' if language_code == 'kk' else 'ru'

        try:
            partner_localization = models.PartnerLocalization.objects.get(partner_id=partner.id,
                                                                          language_id=language_code)
        except ObjectDoesNotExist:
            partner_localization = models.PartnerLocalization.objects.get(partner_id=partner.id,
                                                                          language_id=second_language_code)

        partner_entity.name = partner_localization.name
        partner_entity.description = partner_localization.description

        for reference in partner.partner_references.select_related('type').iterator():
            setattr(partner_entity, reference.type.uid, reference.link)

        try:
            file = models.FileLink.objects.select_related('file').get(
                record_id=partner.id,
                file_class__book__book_class='classes',
                file_class__uid='partners',
                type__book__book_class='file_types',
                type__uid='partner_picture'
            ).file

            partner_entity.picture_uri = f'https://storage.kasipodaq.org/{file.hash}.{file.extension}'
        except ObjectDoesNotExist:
            pass

        return partner_entity

    def get_list(self, partner_entity):
        if partner_entity.search is not None:
            partners = models.Partner.objects \
                .prefetch_related('partner_references') \
                .filter(name__icontains=partner_entity.search)

        else:
            partners = models.Partner.objects \
                .prefetch_related('partner_references') \
                .all()

        partners = partners.filter(**getattr(partner_entity, 'filter_by', {}))
        partners = partners.order_by(partner_entity.order_by)

        paginator = Paginator(partners, partner_entity.count)
        partners = paginator.page(partner_entity.page_number)

        partner_entity_list = []

        language_code = self.context['language'] if 'language' in self.context else 'ru'

        for partner in partners:
            partner_entity_list.append(self.convert(partner, language_code))

        return {
            'partner_list': partner_entity_list,
            'page_number': partner_entity.page_number,
            'count': partner_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def get_by_id(self, resource_id):
        try:
            partner = models.Partner.objects \
                .prefetch_related('partner_references') \
                .get(pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('partner_not_found')

        language_code = self.context['language'] if 'language' in self.context else 'ru'

        return self.convert(partner, language_code)

    def create(self, partner_entity):
        try:
            partner_category = models.Record.objects.get(pk=partner_entity.category_id)
        except ObjectDoesNotExist:
            raise self.make_exception('record_not_found')

        partner = models.Partner()
        partner.category = partner_category

        try:
            file = models.File.objects.get(pk=partner_entity.picture_id)
        except ObjectDoesNotExist:
            raise self.make_exception('file_not_found')

        partner.save()

        for localization in partner_entity.localizations:
            partner_localization = models.PartnerLocalization()
            partner_localization.partner = partner
            partner_localization.language_id = localization['language_id']
            partner_localization.name = localization['name']
            partner_localization.description = localization['description']

            partner_localization.save()

        for social_name in partner_entity.socials:
            partner_reference = models.PartnerReference()
            partner_reference.partner = partner
            partner_reference.type = self.get_book_record('social_networks', social_name)
            partner_reference.link = partner_entity.socials.get(social_name)

            partner_reference.save()

        file_link = models.FileLink()
        file_link.file = file
        file_link.type = self.get_book_record('file_types', 'partner_picture')
        file_link.file_class = self.get_book_record('classes', 'partners')
        file_link.record_id = partner.id

        file_link.save()

        return partner.id

    def edit(self, partner_entity):
        try:
            partner = models.Partner.objects.get(pk=partner_entity.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('partner_not_found')

        for localization in partner_entity.localizations:
            try:
                partner_localization = models.PartnerLocalization.objects.get(partner_id=partner.id,
                                                                              language_id=localization['language_id'])
            except ObjectDoesNotExist:
                partner_localization = models.PartnerLocalization()
                partner_localization.partner = partner
                partner_localization.language_id = localization['language_id']

            partner_localization.name = localization['name']
            partner_localization.description = localization['description']

            partner_localization.save()

        for key in partner_entity.__dict__.keys():
            if hasattr(partner, key):
                setattr(partner, key, getattr(partner_entity, key))

        if hasattr(partner_entity, 'picture_id'):
            try:
                file = models.File.objects.get(pk=partner_entity.picture_id)
            except ObjectDoesNotExist:
                raise self.make_exception('file_not_found')

            models.FileLink.objects.filter(file_class=self.get_book_record('classes', 'partners'),
                                           type=self.get_book_record('file_types', 'partner_picture'),
                                           record_id=partner.id).update(status=0)

            file_link = models.FileLink()
            file_link.file = file
            file_link.type = self.get_book_record('file_types', 'partner_picture')
            file_link.file_class = self.get_book_record('classes', 'partners')
            file_link.record_id = partner.id

            file_link.save()

        partner.save()

        for social_name in partner_entity.socials:
            try:
                partner_reference = models.PartnerReference.objects.get(partner=partner, type__uid=social_name)
                partner_reference.link = partner_entity.socials.get(social_name)
            except ObjectDoesNotExist:
                partner_reference = models.PartnerReference()
                partner_reference.partner = partner
                partner_reference.type = self.get_book_record('social_networks', social_name)
                partner_reference.link = partner_entity.socials.get(social_name)

            partner_reference.save()

    def delete(self, resource_id):
        try:
            partner = models.Partner.objects.get(pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('partner_not_found')

        partner.delete()


class LocalizationGateway(Gateway):
    def get(self, key, language):
        try:
            localization = models.Localization.objects.get(record_id=150,
                                                           attribute_class_id=2,
                                                           language_id=language,
                                                           key=key).value
        except ObjectDoesNotExist:
            localization = key

        return localization

    def list(self, language):
        localizations = dict(models.Localization.objects.filter(record_id=150,
                                                                attribute_class_id=2,
                                                                language_id=language).values_list('key', 'value'))

        return localizations


class FAQGateway(Gateway):
    def __convert(self, faq):
        faq_entity = entities.FAQ()
        faq_entity.resource_id = faq.id
        faq_entity.question = faq.question
        faq_entity.answer = faq.answer
        faq_entity.created_date = faq.created_date
        faq_entity.updated_date = faq.updated_date

        return faq_entity

    def get_list(self, faq_entity):

        faq_entity_list = models.FAQ.objects.all()

        faq_entity_list = faq_entity_list.filter(**getattr(faq_entity, 'filter_by', {}))
        faq_entity_list = faq_entity_list.order_by(faq_entity.order_by)

        paginator = Paginator(faq_entity_list, faq_entity.count)
        faq_entity_list = paginator.page(faq_entity.page_number)

        faq_list = []

        for faq in faq_entity_list:
            faq_list.append(self.__convert(faq))

        return {
            'faq_list': faq_list,
            'page_number': faq_entity.page_number,
            'count': faq_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def get_by_id(self, resource_id):
        try:
            faq = models.FAQ.objects.get(pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('question_not_found')

        return self.__convert(faq)

    def create(self, faq_entity):
        faq = models.FAQ()
        faq.question = faq_entity.question
        faq.answer = faq_entity.answer

        faq.save()

        return faq.id

    def edit(self, faq_entity):
        try:
            faq = models.FAQ.objects.get(pk=faq_entity.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('question_not_found')

        for key in faq_entity.__dict__.keys():
            if hasattr(faq, key):
                setattr(faq, key, getattr(faq_entity, key))

        faq.save()

    def delete(self, resource_id):
        try:
            models.FAQ.objects.get(pk=resource_id).delete()
        except ObjectDoesNotExist:
            raise self.make_exception('question_not_found')


class RecordGateway(Gateway):
    def convert(self, record, language_code='ru'):

        record_class = self.get_book_record('classes', 'record')

        second_language_code = 'ru' if language_code == 'kk' else 'ru'

        try:
            record_localization = models.Localization.objects.get(record_id=record.id,
                                                                  attribute_class=record_class,
                                                                  key='name',
                                                                  language__code=language_code)

        except ObjectDoesNotExist:
            record_localization = models.Localization.objects.get(record_id=record.id,
                                                                  attribute_class=record_class,
                                                                  key='name',
                                                                  language__code=second_language_code)

        record_entity = entities.Record()
        record_entity.resource_id = record.id
        record_entity.name = record_localization.value
        record_entity.sort = record.sort

        return record_entity

    def get_list(self, record_entity):
        book = self.get_book(record_entity.book)

        record_entity_list = models.Record.objects.filter(book=book).order_by('sort')

        record_entity_list = record_entity_list.filter(**getattr(record_entity, 'filter_by', {}))
        record_entity_list = record_entity_list.order_by(record_entity.order_by)

        paginator = Paginator(record_entity_list, record_entity.count)
        record_entity_list = paginator.page(record_entity.page_number)

        record_list = []

        language_code = self.context['language'] if 'language' in self.context else 'ru'

        for record in record_entity_list:
            record_list.append(self.convert(record, language_code))

        return {
            'record_list': record_list,
            'page_number': record_entity.page_number,
            'count': record_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def get_by_id(self, resource_id):
        try:
            record = models.Record.objects.get(pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('record_not_found')

        language_code = self.context['language'] if 'language' in self.context else 'ru'

        return self.convert(record, language_code)

    def create(self, record_entity):
        try:
            book = models.Book.objects.get(book_class=record_entity.book)

        except ObjectDoesNotExist:
            raise self.make_exception('record_not_found')

        record = models.Record()
        record.book = book

        if hasattr(record_entity, 'sort'):
            record.sort = record_entity.sort

        record.save()

        try:
            attribute_class = self.get_book_record('classes', 'record')

        except ObjectDoesNotExist:
            raise self.make_exception('record_not_found')

        for localization in record_entity.localizations:
            record_localization = models.Localization()
            record_localization.attribute_class = attribute_class
            record_localization.record_id = record.id
            record_localization.language_id = localization['language_id']
            record_localization.key = 'name'
            record_localization.value = localization['name']

            record_localization.save()

        return record.id

    def edit(self, record_entity):
        try:
            record = models.Record.objects.get(pk=record_entity.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('record_not_found')

        try:
            attribute_class = self.get_book_record('classes', 'record')

        except ObjectDoesNotExist:
            raise self.make_exception('record_not_found')

        for localization in record_entity.localizations:
            try:
                record_localization = models.Localization.objects.get(record_id=record.id,
                                                                      attribute_class=attribute_class,
                                                                      key='name',
                                                                      language__code='ru')

            except ObjectDoesNotExist:
                record_localization = models.Localization()
                record_localization.attribute_class = attribute_class
                record_localization.record_id = record.id
                record_localization.language_id = localization['language_id']
                record_localization.key = 'name'

            record_localization.value = localization['name']

            record_localization.save()

        for key in record_entity.__dict__.keys():
            if hasattr(record, key):
                setattr(record, key, getattr(record_entity, key))

        record.save()

    def delete(self, resource_id):
        try:
            models.Localization.objects.get(record_id=resource_id,
                                            attribute_class=self.get_book_record('classes', 'record'),
                                            key='name',
                                            language__code='ru').delete()

            models.Record.objects.get(pk=resource_id).delete()
        except ObjectDoesNotExist:
            raise self.make_exception('record_not_found')


class OrderGateway(Gateway):
    def __convert(self, order, person_gateway, union_gateway, article_gateway):

        order_entity = entities.Order()

        order_entity.resource_id = order.id
        order_entity.union = union_gateway.convert(order.union, person_gateway, article_gateway)
        order_entity.person = person_gateway.convert(order.person, models.User.objects.get(person=order.person))
        order_entity.title = order.title
        order_entity.created_date = order.created_date
        order_entity.updated_date = order.updated_date

        try:
            file_links = models.FileLink.objects.filter(file_class=self.get_book_record('classes', 'orders'),
                                                        type=self.get_book_record('file_types', 'orders_file'),
                                                        record_id=order.id)

            files_list = []

            for link in file_links:
                files_list.append({
                    'name': link.file.name,
                    'uri': f'https://storage.kasipodaq.org/{link.file.hash}.{link.file.extension}'
                })

            order_entity.files = files_list

        except ObjectDoesNotExist:
            pass

        return order_entity

    def get_list(self, order_entity, union_gateway, person_gateway, article_gateway):

        if order_entity.parent_orders:
            union = models.Union.objects.get(pk=order_entity.union.resource_id)
            order_entity.union.resource_id = union.parent_id

        if order_entity.all:
            union = models.Union.objects.get(pk=order_entity.union.resource_id)
            union_list = [union.id]

            if hasattr(union, 'association_id') and union.association_id is not None:
                union_list.append(union.association_id)

            while union.parent is not None:
                union = union.parent
                union_list.append(union.id)

            order_entity_list = models.Order.objects.filter(union__id__in=union_list, status=1)

        else:
            order_entity_list = models.Order.objects.filter(union__id=order_entity.union.resource_id, status=1)

        if order_entity.search is not None:
            order_entity_list = order_entity_list.filter(title__icontains=order_entity.search)

        order_entity_list = order_entity_list.filter(**getattr(order_entity, 'filter_by', {}))
        order_entity_list = order_entity_list.order_by(order_entity.order_by)

        paginator = Paginator(order_entity_list, order_entity.count)
        order_entity_list = paginator.page(order_entity.page_number)

        order_list = []

        for order in order_entity_list:
            order_list.append(self.__convert(order, person_gateway, union_gateway, article_gateway))

        return {
            'order_list': order_list,
            'page_number': order_entity.page_number,
            'count': order_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def get_by_id(self, union_gateway, person_gateway, article_gateway, resource_id):
        try:
            order = models.Order.objects.get(pk=resource_id, status=1)
        except ObjectDoesNotExist:
            raise self.make_exception('order_not_found')

        return self.__convert(order, person_gateway, union_gateway, article_gateway)

    def create(self, order_entity):
        try:
            union_member = models.UnionMember.objects.get(Q(status=1) | Q(status=101),
                                                          person__id=order_entity.person.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('union_member_not_found')

        order = models.Order()
        order.title = order_entity.title
        order.person = models.Person.objects.get(pk=order_entity.person.resource_id)
        order.union = union_member.union

        order.save()

        try:
            models.FileLink.objects.filter(file_class=self.get_book_record('classes', 'orders'),
                                           type=self.get_book_record('file_types', 'orders_file'),
                                           record_id=order.id).delete()

            for file in order_entity.files:
                file = models.File.objects.get(pk=file.resource_id)

                file_link = models.FileLink()
                file_link.file = file
                file_link.type = self.get_book_record('file_types', 'orders_file')
                file_link.file_class = self.get_book_record('classes', 'orders')
                file_link.record_id = order.id

                file_link.save()
        except ObjectDoesNotExist:
            pass

        notify_id = self.notify(order.person, self.get_book_record('classes', 'orders'), order.id,
                                f'Опубликован приказ "{order.title}"')

        channel_layer = get_channel_layer()

        event = {
            'event': 'EVENT_NEW_ORDER',
            'message': 'Опубликован приказ "' + order.title + '"',
            'record_id': order.id,
            'resource_id': notify_id
        }

        async_to_sync(channel_layer.group_send)('union_' + str(union_member.union.id), {
            'type': 'send_event',
            'event_type': json.dumps(event)
        })

        return order.id

    def edit(self, order_entity):
        try:
            order = models.Order.objects.get(pk=order_entity.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('order_not_found')

        for key in order_entity.__dict__.keys():
            if hasattr(order, key):
                setattr(order, key, getattr(order_entity, key))

        order.save()

        try:
            if hasattr(order_entity, 'files'):
                models.FileLink.objects.filter(file_class=self.get_book_record('classes', 'orders'),
                                               type=self.get_book_record('file_types', 'orders_file'),
                                               record_id=order.id).delete()

                for file in order_entity.files:
                    file = models.File.objects.get(pk=file.resource_id)

                    file_link = models.FileLink()
                    file_link.file = file
                    file_link.type = self.get_book_record('file_types', 'orders_file')
                    file_link.file_class = self.get_book_record('classes', 'orders')
                    file_link.record_id = order.id

                    file_link.save()
        except ObjectDoesNotExist:
            pass

    def delete(self, resource_id):
        try:
            order = models.Order.objects.get(status=1, pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('order_not_found')

        order.status = 2
        order.save()


class NotificationGateway(Gateway):
    def convert(self, notification, read_notification):
        notification_entity = entities.Notification()
        notification_entity.resource_id = notification.id
        notification_entity.record_id = notification.record_id
        notification_entity.notification_class = notification.attribute_class.uid
        notification_entity.content = notification.content

        if notification.id in read_notification:
            notification_entity.is_seen = True

        else:
            notification_entity.is_seen = False

        return notification_entity

    def get_list(self, notification_entity, person_id):
        try:
            union = models.UnionMember.objects.get(Q(status=1) | Q(status=101), person__id=person_id).union
            union_list = [union.id]

            if hasattr(union, 'association_id') and union.association_id is not None:
                union_list.append(union.association_id)

            while union.parent is not None:
                union = union.parent
                union_list.append(union.id)

        except ObjectDoesNotExist:
            union_list = []

        notification_list = models.Notification.objects.filter(
            Q(union__id__in=union_list) | Q(person__id=person_id)).order_by('-created_date')

        try:
            read_notification = models.PersonNotification.objects.get(person__id=person_id).read_notification
        except ObjectDoesNotExist:
            read_notification = []

        notification_list = notification_list.filter(**getattr(notification_entity, 'filter_by', {}))
        notification_list = notification_list.order_by(notification_entity.order_by)

        paginator = Paginator(notification_list, notification_entity.count)
        notification_list = paginator.page(notification_entity.page_number)

        notification_list = [self.convert(notification, read_notification) for notification in notification_list]

        return {
            'notification_list': notification_list,
            'page_number': notification_entity.page_number,
            'count': notification_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def read_notification(self, notification_id, person_id):
        try:
            person_notification = models.PersonNotification.objects.get(person__id=person_id)

        except ObjectDoesNotExist:
            person_notification = models.PersonNotification()
            person_notification.person = models.Person.objects.get(pk=person_id)

        person_notification.read_notification.append(notification_id)
        person_notification.save()


class RuleGateway(Gateway):
    def convert(self, rule):
        rule_entity = entities.Rule()
        rule_entity.use_case = rule.use_case.uid
        rule_entity.method = rule.method.uid
        rule_entity.rules = json.loads(rule.rules)

        return rule_entity

    def get_list(self, use_case, method):
        try:
            rules = models.Rule.objects.filter(use_case=self.get_book_record('use_case', use_case),
                                               method=self.get_book_record('use_case_methods', method))
        except ObjectDoesNotExist:
            return []

        return [json.loads(rule.rules) for rule in rules]

    def get_permission_list(self):
        rules = models.Rule.objects.all()

        return [self.convert(rule) for rule in rules]


class RevisionGateway(Gateway):
    def convert(self, revision, person_gateway, union_gateway, article_gateway, record_gateway, person_id=None,
                language_code='ru'):

        revision_entity = entities.Revision()

        revision_entity.resource_id = revision.id
        revision_entity.union = union_gateway.convert(revision.union, person_gateway, article_gateway)
        revision_entity.person = person_gateway.convert(revision.person,
                                                        models.User.objects.get(person=revision.person))
        revision_entity.type = record_gateway.convert(revision.type)
        revision_entity.is_person_answered = False
        revision_entity.start_date = revision.start_date.strftime("%Y-%m-%d %H:%M:%S")
        revision_entity.finish_date = revision.finish_date.strftime("%Y-%m-%d %H:%M:%S")
        revision_entity.status = revision.status

        try:
            if person_id is not None:
                models.RevisionStatistics.objects.get(revision=revision, person_id=person_id)
        except ObjectDoesNotExist:
            pass

        if hasattr(revision, 'percent_threshold'):
            revision_entity.percent_threshold = int(revision.percent_threshold * 100)

        try:
            revision_localization = models.RevisionLocalization.objects.get(record_id=revision.id,
                                                                            attribute_class__book__book_class='classes',
                                                                            attribute_class__uid='revisions',
                                                                            key='name',
                                                                            language__code='ru')

            revision_entity.name = revision_localization.value

        except ObjectDoesNotExist:
            raise self.make_exception('record_not_found')

        try:
            revision_localization = models.RevisionLocalization.objects.get(record_id=revision.id,
                                                                            attribute_class__book__book_class='classes',
                                                                            attribute_class__uid='revisions',
                                                                            key='decree',
                                                                            language__code='ru')

            revision_entity.decree = revision_localization.value

        except ObjectDoesNotExist:
            pass

        return revision_entity

    def get_list(self, revision_entity, person_gateway, union_gateway, article_gateway, record_gateway):
        union = models.Union.objects.get(pk=revision_entity.filter_by['union_id'])

        union_ids = [union.id]

        if hasattr(union, 'association_id') and union.association_id is not None:
            union_ids.append(union.association_id)

        while union.parent is not None:
            union = union.parent
            union_ids.append(union.id)

        if revision_entity.is_archive:
            revision_entity_list = models.Revision.objects.filter(~Q(status=2),
                                                                  finish_date__lte=datetime.datetime.now())
            revision_entity_list = revision_entity_list.filter(**getattr(revision_entity, 'filter_by', {}))

        elif hasattr(revision_entity, 'is_answered') and revision_entity.is_answered:
            revision_ids = list(
                models.RevisionStatistics.objects.filter(person__id=revision_entity.person.resource_id).values_list(
                    'revision_id', flat=True))
            revision_entity_list = models.Revision.objects.filter(status=1, id__in=revision_ids,
                                                                  union__id__in=union_ids)

            del revision_entity.filter_by['union_id']

        elif hasattr(revision_entity, 'is_answered') and not revision_entity.is_answered:
            revision_ids = list(
                models.RevisionStatistics.objects.filter(person__id=revision_entity.person.resource_id).values_list(
                    'revision_id', flat=True))

            revision_entity_list = models.Revision.objects.filter(~Q(id__in=revision_ids), status=1,
                                                                  finish_date__gte=datetime.datetime.now(),
                                                                  union__id__in=union_ids)

            del revision_entity.filter_by['union_id']

        else:
            revision_entity_list = models.Revision.objects.filter(~Q(status=2),
                                                                  finish_date__gt=datetime.datetime.now())
            revision_entity_list = revision_entity_list.filter(**getattr(revision_entity, 'filter_by', {}))

        revision_entity_list = revision_entity_list.order_by(revision_entity.order_by)

        paginator = Paginator(revision_entity_list, revision_entity.count)
        revision_entity_list = paginator.page(revision_entity.page_number)

        revision_list = []

        for revision in revision_entity_list:
            revision_list.append(
                self.convert(revision, person_gateway, union_gateway, article_gateway, record_gateway,
                             revision_entity.person.resource_id))

        return {
            'revision_list': revision_list,
            'page_number': revision_entity.page_number,
            'count': revision_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def get_by_id(self, person_gateway, union_gateway, article_gateway, record_gateway, resource_id):
        try:
            revision = models.Revision.objects.get(pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('revision_not_found')

        return self.convert(revision, person_gateway, union_gateway, article_gateway, record_gateway)

    def create(self, revision_entity):
        try:
            union_member = models.UnionMember.objects.get(status=101, person__id=revision_entity.person.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('union_member_not_found')

        self.get_book_record_by_id('revision_category', revision_entity.type.resource_id)

        revision = models.Revision()
        revision.person = models.Person.objects.get(pk=revision_entity.person.resource_id)
        revision.union = union_member.union
        revision.type = models.Record.objects.get(pk=revision_entity.type.resource_id)
        revision.start_date = revision_entity.start_date
        revision.finish_date = revision_entity.finish_date

        if hasattr(revision_entity, 'percent_threshold'):
            revision.percent_threshold = float(revision_entity.percent_threshold) / 100

        revision.save()

        attribute_class = self.get_book_record('classes', 'revisions')

        revisions_localization = models.RevisionLocalization()
        revisions_localization.attribute_class = attribute_class
        revisions_localization.record_id = revision.id
        revisions_localization.language = models.Language.objects.get(code='ru')
        revisions_localization.key = 'name'
        revisions_localization.value = revision_entity.name

        revisions_localization.save()

        return revision.id

    def edit(self, revision_entity):
        try:
            revision = models.Revision.objects.get(pk=revision_entity.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('revision_not_found')

        localization_class = self.get_book_record('classes', 'revisions')

        if revision.status != 0:
            if hasattr(revision_entity, 'decree'):
                revision_localization = models.RevisionLocalization()
                revision_localization.attribute_class = localization_class
                revision_localization.record_id = revision.id
                revision_localization.language = models.Language.objects.get(code='ru')
                revision_localization.key = 'decree'
                revision_localization.value = revision_entity.decree

                revision_localization.save()

            else:
                raise self.make_exception('can_not_change_publish_revision')

        else:

            for key in revision_entity.__dict__.keys():
                if hasattr(revision, key):
                    if key == 'percent_threshold':
                        revision.percent_threshold = float(revision_entity.percent_threshold) / 100
                    else:
                        setattr(revision, key, getattr(revision_entity, key))

                try:
                    revision_localization = models.RevisionLocalization.objects.get(record_id=revision.id,
                                                                                    key=key,
                                                                                    attribute_class=localization_class,
                                                                                    language__code='ru')

                    revision_localization.value = getattr(revision_entity, key)
                    revision_localization.save()

                except ObjectDoesNotExist:
                    pass

            revision.save()

    def delete(self, resource_id):
        try:
            revision = models.Revision.objects.get(pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('revision_not_found')

        revision.status = 2
        revision.save()

    def publish(self, resource_id):
        try:
            revision = models.Revision.objects.get(status=0, pk=resource_id)
            localization_class = self.get_book_record('classes', 'revisions')
            revision_title = models.RevisionLocalization.objects.get(record_id=revision.id,
                                                                     key='name',
                                                                     attribute_class=localization_class,
                                                                     language__code='ru').value
        except ObjectDoesNotExist:
            raise self.make_exception('revision_not_found')

        questions = models.Question.objects.filter(revision=revision)
        questions_ids = list(questions.values_list('id', flat=True))

        if len(questions_ids) == 0:
            raise self.make_exception('question_not_found')

        type_test = self.get_book_record('revision_category', 'test')

        if revision.type == type_test:
            for question_id in questions_ids:
                has_valid_answer = models.Answer.objects.filter(question__id=question_id, is_right=True).exists()
                if not has_valid_answer:
                    raise self.make_exception('need_one_right_answer')

        else:
            for question_id in questions_ids:
                count_answer = models.Answer.objects.filter(question__id=question_id).count()
                if count_answer < 2:
                    raise self.make_exception('need_two_answer')

        if revision.status == 0:
            revision.status = 1
            revision.save()

            if revision.type == type_test:
                text = f'Опубликован новый тест "{revision_title}"'

            else:
                text = f'Опубликован новый опрос "{revision_title}"'

            notify_id = self.notify(revision.person, self.get_book_record('classes', 'revisions'), revision.id,
                                    text)

            channel_layer = get_channel_layer()

            event = {
                'event': 'EVENT_NEW_REVISION',
                'message': text,
                'record_id': revision.id,
                'resource_id': notify_id
            }

            async_to_sync(channel_layer.group_send)('union_' + str(revision.union.id), {
                'type': 'send_event',
                'event_type': json.dumps(event)
            })

        else:
            raise self.make_exception('can_not_change_publish_revision')

    def finish(self, revision_id, person_id):
        try:
            revision = models.Revision.objects.get(pk=revision_id)
        except ObjectDoesNotExist:
            raise self.make_exception('revision_not_found')

        # if revision.status == 0:
        #     raise Exception('Нельзя завершить еще не опубликованный тест')

        try:
            person = models.Person.objects.get(pk=person_id)
        except ObjectDoesNotExist:
            raise self.make_exception('person_not_found')

        try:
            models.RevisionStatistics.objects.get(person__id=person.id, revision__id=revision.id).delete()

        except ObjectDoesNotExist:
            pass

        question_ids = list(models.Question.objects.filter(revision=revision).values_list('id', flat=True))
        total_questions = len(question_ids)

        answer_ids = list(models.Answer.objects.filter(question__id__in=question_ids).values_list('id', flat=True))
        total_person_answers = list(models.PersonAnswer.objects.filter(answer__id__in=answer_ids, person=person))
        total_person_answers = len(total_person_answers)

        if total_person_answers < total_questions:
            raise self.make_exception('you_need_to_finish_revision')

        type_test = self.get_book_record('revision_category', 'test')

        if revision.type == type_test:
            valid_answer_ids = models.Answer.objects.filter(question__id__in=question_ids,
                                                            is_right=True).values_list(
                'id')

        else:
            valid_answer_ids = models.Answer.objects.filter(question__id__in=question_ids).values_list('id')

        valid_person_answers = len(models.PersonAnswer.objects.filter(answer__id__in=valid_answer_ids,
                                                                      person=person))

        revision_statistic = models.RevisionStatistics()
        revision_statistic.revision = revision
        revision_statistic.person = person
        revision_statistic.total_questions = total_questions
        revision_statistic.valid_answers = valid_person_answers
        revision_statistic.is_passed = False

        if revision.percent_threshold < (float(valid_person_answers) / float(total_questions)):
            revision_statistic.is_passed = True

        revision_statistic.save()

        return {
            'total_questions': total_questions,
            'valid_answers': valid_person_answers,
            'invalid_answers': total_questions - valid_person_answers,
            'is_passed': revision_statistic.is_passed,
        }

    def union_members(self, revision_id):
        try:
            revision = models.Revision.objects.get(pk=revision_id, status=1)
        except ObjectDoesNotExist:
            raise self.make_exception('revision_not_found')

        union_ids = [revision.union.id]

        current_union_ids = [revision.union.id]

        while True:
            current_union_ids = list(
                models.Union.objects.filter(parent__id__in=current_union_ids, status=1).values_list('id',
                                                                                                    flat=True))

            if len(current_union_ids) == 0:
                break
            else:
                union_ids = union_ids + current_union_ids

        person_ids = models.UnionMember.objects.filter((Q(join_date__lt=revision.start_date) & (
                Q(leave_date__gt=revision.start_date) | Q(leave_date__isnull=True)) |
                                                        Q(join_date__lt=revision.finish_date) & (
                                                                Q(leave_date__gt=revision.finish_date) | Q(
                                                            leave_date__isnull=True))) &
                                                       Q(union__id__in=union_ids)).values_list('person_id',
                                                                                               flat=True).distinct(
            'person_id')

        return list(person_ids)

    def total_finished(self, revision_id, person_ids):
        total_finished = models.RevisionStatistics.objects.filter(revision__id=revision_id,
                                                                  person__id__in=person_ids).count()

        return total_finished

    def total_complete(self, revision_id, person_ids):
        total_complete = models.RevisionStatistics.objects.filter(revision__id=revision_id,
                                                                  person__id__in=person_ids,
                                                                  is_passed=True).count()

        return total_complete

    def person_test_statistic(self, revision_id, person_id):
        try:
            revision_person_statistic = models.RevisionStatistics.objects.get(person__id=person_id,
                                                                              revision__id=revision_id)
        except ObjectDoesNotExist:
            raise self.make_exception('statistics_not_found')

        return {
            'total_questions': revision_person_statistic.total_questions,
            'valid_answers': revision_person_statistic.valid_answers,
            'invalid_answers': revision_person_statistic.total_questions - revision_person_statistic.valid_answers,
            'is_passed': revision_person_statistic.is_passed,
        }

    def vote_statistic(self, revision_id, question_gateway, answer_gateway, union_gateway, person_gateway,
                       article_gateway, record_gateway):
        try:
            revision = models.Revision.objects.get(pk=revision_id)
        except ObjectDoesNotExist:
            raise self.make_exception('revision_not_found')

        union_members = self.union_members(revision_id)
        total_finished = self.total_finished(revision_id, union_members)

        revision_entity = self.convert(revision, person_gateway, union_gateway, article_gateway, record_gateway)

        revision_entity.total_union_members = len(union_members)
        revision_entity.total_finished = total_finished
        revision_entity.not_finished = revision_entity.total_union_members - revision_entity.total_finished

        questions = models.Question.objects.filter(revision=revision)

        revision_entity.questions = []

        for question in questions:
            answer_list = models.Answer.objects.filter(question=question)
            revision_entity.questions.append(
                question_gateway.convert(question, answer_gateway, self, person_gateway, union_gateway,
                                         article_gateway,
                                         record_gateway, answer_list))

        return revision_entity

    def revision_members(self, revision_id, person_gateway):
        try:
            models.Revision.objects.get(pk=revision_id)
        except ObjectDoesNotExist:
            raise self.make_exception('revision_not_found')

        person_ids = list(
            models.RevisionStatistics.objects.filter(revision__id=revision_id).values_list(
                'person_id', flat=True))

        persons = models.Person.objects.filter(id__in=person_ids)

        person_list = []

        for person in persons:
            person_list.append(
                person_gateway.convert(person, models.User.objects.get(person__id=person.id, status__in=[1, 101])))

        return person_list


class QuestionGateway(Gateway):
    def convert(self, question, answer_gateway, revision_gateway, person_gateway, union_gateway, article_gateway,
                record_gateway, answer_list=None):

        question_entity = entities.Question()
        question_entity.resource_id = question.id
        question_entity.revision = revision_gateway.convert(
            models.Revision.objects.get(pk=question.revision.id), person_gateway, union_gateway,
            article_gateway, record_gateway)
        question_entity.has_alternate_answer = question.has_alternate_answer

        try:
            question_localization = models.RevisionLocalization.objects.get(record_id=question.id,
                                                                            attribute_class__book__book_class='classes',
                                                                            attribute_class__uid='questions',
                                                                            key='question',
                                                                            language__code='ru')

            question_entity.question = question_localization.value

        except ObjectDoesNotExist:
            raise self.make_exception('record_not_found')

        if answer_list is not None:
            question_entity.answers = []

            for answer in answer_list:
                question_entity.answers.append(answer_gateway.convert(answer, person_gateway, None))

        return question_entity

    def get_list(self, question_entity, answer_gateway, revision_gateway, person_gateway, union_gateway,
                 article_gateway, record_gateway):
        question_entity_list = models.Question.objects.filter(
            revision__id=question_entity.revision.resource_id)

        question_entity_list = question_entity_list.filter(**getattr(question_entity, 'filter_by', {}))
        question_entity_list = question_entity_list.order_by(question_entity.order_by)

        paginator = Paginator(question_entity_list, question_entity.count)
        question_entity_list = paginator.page(question_entity.page_number)

        question_list = []

        for question in question_entity_list:
            question_list.append(
                self.convert(question, answer_gateway, revision_gateway, person_gateway, union_gateway,
                             article_gateway,
                             record_gateway))

        return {
            'question_list': question_list,
            'page_number': question_entity.page_number,
            'count': question_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def get_by_id(self, answer_gateway, revision_gateway, person_gateway, union_gateway, article_gateway,
                  record_gateway, resource_id, person_id=None):
        try:
            question = models.Question.objects.get(pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('question_not_found')

        if person_id is not None:
            answer_list = models.Answer.objects.filter(Q(person__id=person_id) | Q(person=None),
                                                       question=question).order_by('person_id')

        else:
            answer_list = models.Answer.objects.filter(person=None, question=question).order_by('?')

        return self.convert(question, answer_gateway, revision_gateway, person_gateway, union_gateway,
                            article_gateway,
                            record_gateway, answer_list)

    def create(self, question_entity):
        try:
            revision = models.Revision.objects.get(pk=question_entity.revision.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('revision_not_found')

        if revision.status != 0:
            raise self.make_exception('can_not_change_publish_revision')

        question = models.Question()
        question.revision = revision

        type_test = self.get_book_record('revision_category', 'test')
        if revision.type == type_test and question_entity.has_alternate_answer:
            raise self.make_exception('cannot_add_alternate_answer')

        else:
            question.has_alternate_answer = question_entity.has_alternate_answer

        question.save()

        attribute_class = self.get_book_record('classes', 'questions')

        question_localization = models.RevisionLocalization()
        question_localization.attribute_class = attribute_class
        question_localization.record_id = question.id
        question_localization.language = models.Language.objects.get(code='ru')
        question_localization.key = 'question'
        question_localization.value = question_entity.question

        question_localization.save()

        return question.id

    def edit(self, question_entity):
        try:
            question = models.Question.objects.get(pk=question_entity.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('question_not_found')

        for key in question_entity.__dict__.keys():
            if key == 'question':
                localization_class = self.get_book_record('classes', 'questions')

                question_localization = models.RevisionLocalization.objects.get(record_id=question.id,
                                                                                key='question',
                                                                                attribute_class=localization_class,
                                                                                language__code='ru')

                question_localization.value = question_entity.question

                question_localization.save()

            elif key == 'has_alternate_answer':
                question.has_alternate_answer = question_entity.has_alternate_answer
                question.save()

    def delete(self, resource_id):
        try:
            question = models.Question.objects.get(pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('question_not_found')

        if question.revision.status == 0:
            try:
                models.Answer.objects.filter(question=question).delete()

            except ObjectDoesNotExist:
                pass

            attribute_class = self.get_book_record('classes', 'questions')

            models.RevisionLocalization.objects.get(record_id=question.id,
                                                    key='question',
                                                    attribute_class=attribute_class,
                                                    language__code='ru').delete()

            question.delete()

        else:
            raise self.make_exception('can_not_change_publish_revision')


class CommentGateway(Gateway):
    def convert(self, comment, person_gateway):
        question_entity = entities.Comment()
        question_entity.resource_id = comment.id
        question_entity.author = person_gateway.convert(comment.person,
                                                        models.User.objects.get(person=comment.person, status=1))
        question_entity.comment = comment.comment

        return question_entity

    def get_list(self, comment_entity, person_gateway):
        comment_entity_list = models.Comment.objects.filter(question_id=comment_entity.question.resource_id)

        comment_entity_list = comment_entity_list.filter(**getattr(comment_entity, 'filter_by', {}))
        comment_entity_list = comment_entity_list.order_by(comment_entity.order_by)

        paginator = Paginator(comment_entity_list, comment_entity.count)
        comment_entity_list = paginator.page(comment_entity.page_number)

        comment_list = []

        for comment in comment_entity_list:
            comment_list.append(self.convert(comment, person_gateway))

        return {
            'comment_list': comment_list,
            'page_number': comment_entity.page_number,
            'count': comment_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def create(self, comment_entity):
        try:
            question = models.Question.objects.get(pk=comment_entity.question.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('question_not_found')

        if question.revision.status == 0:
            raise self.make_exception('can_not_answer_not_publish_revision')

        comment = models.Comment()
        comment.person = models.Person.objects.get(pk=comment_entity.author.resource_id)
        comment.question = question
        comment.comment = comment_entity.comment

        comment.save()

        return comment.id


class AnswerGateway(Gateway):
    def convert(self, answer, person_gateway, person_id):
        answer_entity = entities.Answer()

        answer_entity.resource_id = answer.id
        answer_entity.total_responded = answer.total_responded
        answer_entity.is_right = answer.is_right

        if person_id is not None:
            try:
                models.PersonAnswer.objects.get(person__id=person_id, answer=answer)
                answer_entity.is_person_answer = True
            except ObjectDoesNotExist:
                answer_entity.is_person_answer = False

        if answer.person is not None:
            answer_entity.answer = answer.alternate_answer
            answer_entity.person = person_gateway.convert(answer.person,
                                                          models.User.objects.get(person=answer.person, status=1))

            return answer_entity

        try:
            answer_localization = models.RevisionLocalization.objects.get(record_id=answer.id,
                                                                          attribute_class__book__book_class='classes',
                                                                          attribute_class__uid='answers',
                                                                          key='answer',
                                                                          language__code='ru')

            answer_entity.answer = answer_localization.value

        except ObjectDoesNotExist:
            raise self.make_exception('record_not_found')

        return answer_entity

    def get_list(self, answer_entity, person_gateway, person_id):
        try:
            question = models.Question.objects.get(pk=answer_entity.question.resource_id)

        except ObjectDoesNotExist:
            raise self.make_exception('question_not_found')

        answer_entity_list = models.Answer.objects.filter(question__id=answer_entity.question.resource_id)

        if question.revision.status == 1:
            answer_entity_list = answer_entity_list.order_by('?')

        answer_entity_list = answer_entity_list.filter(**getattr(answer_entity, 'filter_by', {}))

        paginator = Paginator(answer_entity_list, answer_entity.count)
        answer_entity_list = paginator.page(answer_entity.page_number)

        answer_list = []

        for answer in answer_entity_list:
            answer_list.append(
                self.convert(answer, person_gateway, person_id))

        return {
            'answer_list': answer_list,
            'page_number': answer_entity.page_number,
            'count': answer_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def create(self, answer_entity):
        try:
            question = models.Question.objects.get(pk=answer_entity.question.resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('question_not_found')

        if question.revision.status != 0 and not hasattr(answer_entity, 'alternate_answer'):
            raise self.make_exception('can_not_change_publish_revision')

        answer = models.Answer()
        answer.question = question

        if hasattr(answer_entity, 'is_right'):
            answer.is_right = answer_entity.is_right

        if hasattr(answer_entity, 'alternate_answer'):
            answer.total_responded = answer.total_responded + 1
            answer.person = models.Person.objects.get(pk=answer_entity.person.resource_id)
            answer.alternate_answer = answer_entity.alternate_answer
            answer.save()

        else:
            answer.save()

            attribute_class = self.get_book_record('classes', 'answers')

            answer_localization = models.RevisionLocalization()
            answer_localization.attribute_class = attribute_class
            answer_localization.record_id = answer.id
            answer_localization.language = models.Language.objects.get(code='ru')
            answer_localization.key = 'answer'
            answer_localization.value = answer_entity.answer

            answer_localization.save()

        return answer.id

    def edit(self, answer_id, answer_entity):
        try:
            answer = models.Answer.objects.get(pk=answer_id)
        except ObjectDoesNotExist:
            raise self.make_exception('answer_not_found')

        if answer.question.revision.status == 1:
            raise self.make_exception('can_not_change_publish_revision')

        localization_class = self.get_book_record('classes', 'answers')

        answer_localization = models.RevisionLocalization.objects.get(record_id=answer_id,
                                                                      key='answer',
                                                                      attribute_class=localization_class,
                                                                      language__code='ru')

        answer_localization.value = answer_entity.answer
        answer_localization.save()

        for key in answer_entity.__dict__.keys():
            if hasattr(answer, key):
                setattr(answer, key, getattr(answer_entity, key))

        answer.save()

    def delete(self, resource_id):
        try:
            answer = models.Answer.objects.get(pk=resource_id)
        except ObjectDoesNotExist:
            raise self.make_exception('answer_not_found')

        if answer.question.revision.status == 0:
            attribute_class = self.get_book_record('classes', 'answers')

            models.RevisionLocalization.objects.get(record_id=answer.id,
                                                    key='answer',
                                                    attribute_class=attribute_class,
                                                    language__code='ru').delete()

            answer.delete()

        else:
            raise self.make_exception('can_not_change_publish_revision')


class PersonAnswerGateway(Gateway):
    def create(self, person_answer_entity, answer_gateway):
        person_answer = models.PersonAnswer()
        person_answer.person = models.Person.objects.get(pk=person_answer_entity.person.resource_id)

        if hasattr(person_answer_entity, 'alternate_answer'):
            try:
                question = models.Question.objects.get(pk=person_answer_entity.question.resource_id)
            except ObjectDoesNotExist:
                raise self.make_exception('question_not_found')

            type_test = self.get_book_record('revision_category', 'test')

            if question.revision.type == type_test or not question.has_alternate_answer:
                raise self.make_exception('question_has_not_alternate_answer')

            answer_entity = entities.Answer()
            answer_entity.question.resource_id = question.id
            answer_entity.person.resource_id = person_answer_entity.person.resource_id
            answer_entity.alternate_answer = person_answer_entity.alternate_answer

            models.Answer.objects.filter(person__id=person_answer_entity.person.resource_id,
                                         question=question).delete()

            answer_id = answer_gateway.create(answer_entity)

            person_answer.answer = models.Answer.objects.get(pk=answer_id)

        else:
            answer = models.Answer.objects.get(pk=person_answer_entity.answer.resource_id)
            question = answer.question
            person_answer.answer = answer

            answer.total_responded += 1
            answer.save()

        person_answers = models.PersonAnswer.objects.filter(person__id=person_answer_entity.person.resource_id,
                                                            answer__question=question)

        for person_answer in person_answers:
            answer = models.Answer.objects.get(pk=person_answer.answer.id)
            answer.total_responded -= 1
            answer.save()

            person_answer.delete()

        person_answer.save()

        return person_answer.id


class ReportGateway(Gateway):
    def __convert(self, report, union_id):

        language_code = self.context['language']

        second_language_code = 'ru' if language_code == 'kk' else 'ru'

        try:
            union_report_id = models.Report.objects.get(union_id=union_id, type_id=report.id).id

            report_file = models.FileLink.objects.prefetch_related('file').get(record_id=union_report_id,
                                                                               file_class__book__book_class='classes',
                                                                               file_class__uid='report',
                                                                               type__book__book_class='file_types',
                                                                               type__uid='report')

        except ObjectDoesNotExist:
            union_report_id = None
            report_file = None

        report_entity = entities.Report()
        report_entity.resource_id = union_report_id
        report_entity.union.resource_id = union_id
        report_entity.date_created = None

        report_entity.type = {
            'resource_id': report.id,
            'type': report.uid
        }

        try:
            report_entity.name = models.Localization.objects.get(record_id=report.id,
                                                                 language_id=language_code,
                                                                 attribute_class__book__book_class='classes',
                                                                 attribute_class__uid='record',
                                                                 key='name').value
        except ObjectDoesNotExist:
            report_entity.name = models.Localization.objects.get(record_id=report.id,
                                                                 language_id=second_language_code,
                                                                 attribute_class__book__book_class='classes',
                                                                 attribute_class__uid='record',
                                                                 key='name').value

        fields = models.Record.objects.filter(book__book_class='reports', parent_id=report.id, status=1).order_by(
            'sort', 'id')

        report_entity.fields = []

        for field in fields:
            try:
                field_name = models.Localization.objects.get(record_id=field.id,
                                                             language_id=language_code,
                                                             attribute_class__book__book_class='classes',
                                                             attribute_class__uid='record',
                                                             key='name').value
            except ObjectDoesNotExist:
                field_name = models.Localization.objects.get(record_id=field.id,
                                                             language_id=second_language_code,
                                                             attribute_class__book__book_class='classes',
                                                             attribute_class__uid='record',
                                                             key='name').value

            report_entity.fields.append({
                'key': field.uid,
                'name': field_name
            })

        if report_file is not None:
            report_entity.file = {
                'name': report_file.file.name,
                'uri': f'https://storage.kasipodaq.org/{report_file.file.hash}.{report_file.file.extension}'
            }
            report_entity.date_created = report_file.created_date

        return report_entity

    def get_list(self, report_entity, person_id):
        if report_entity.union.resource_id is None:
            try:
                report_entity.union.resource_id = models.UnionMember.objects.get(person_id=person_id,
                                                                                 status__in=[101]).union_id

            except ObjectDoesNotExist:
                raise self.make_exception('union_not_found')

        reports = models.Record.objects.filter(book__book_class='reports', parent_id=None, status=1).order_by('sort')

        paginator = Paginator(reports, report_entity.count)
        reports = paginator.page(report_entity.page_number)

        report_list = []

        for report in reports:
            report_list.append(self.__convert(report, report_entity.union.resource_id))

        return {
            'report_list': report_list,
            'page_number': report_entity.page_number,
            'count': report_entity.count,
            'total_count': paginator.count,
            'page_count': paginator.num_pages
        }

    def create(self, report_entity, file_gateway, person_id=None):
        if person_id is None:
            try:
                report = models.Report.objects.get(pk=report_entity.resource_id)
                report_entity.union.resource_id = report.union_id

            except ObjectDoesNotExist:
                raise self.make_exception('report_not_found')

        else:
            if report_entity.union.resource_id is None:
                try:
                    report_entity.union.resource_id = models.UnionMember.objects.get(person_id=person_id,
                                                                                     status__in=[101]).union_id

                except ObjectDoesNotExist:
                    raise self.make_exception('union_not_found')

            try:
                report = models.Report.objects.get(union_id=report_entity.union.resource_id,
                                                   type_id=report_entity.type.resource_id)

            except ObjectDoesNotExist:
                report = models.Report()
                report.union_id = report_entity.union.resource_id
                report.type_id = report_entity.type.resource_id
                report.save()

        type_record = models.Record.objects.get(pk=report.type_id)

        language_code = self.context['language']

        second_language_code = 'ru' if language_code == 'kk' else 'ru'

        try:
            type_name = models.Localization.objects.get(record_id=type_record.id,
                                                        language_id=language_code,
                                                        attribute_class=self.get_book_record('classes', 'record'),
                                                        key='name').value
        except ObjectDoesNotExist:
            type_name = models.Localization.objects.get(record_id=type_record.id,
                                                        language_id=second_language_code,
                                                        attribute_class=self.get_book_record('classes', 'record'),
                                                        key='name').value

        report_entity.type.type = type_record.uid
        report_entity.type.name = type_name

        full_fields = {}

        for field in report_entity.fields:
            try:
                field = models.Record.objects.get(uid=field, parent_id=report.type_id,
                                                  book__book_class='reports')

                try:
                    field_name = models.Localization.objects.get(record_id=type_record.id,
                                                                 language_id=language_code,
                                                                 attribute_class=self.get_book_record('classes',
                                                                                                      'record'),
                                                                 key='name').value
                except ObjectDoesNotExist:
                    field_name = models.Localization.objects.get(record_id=type_record.id,
                                                                 language_id=second_language_code,
                                                                 attribute_class=self.get_book_record('classes',
                                                                                                      'record'),
                                                                 key='name').value

                full_fields[field.uid] = field_name

            except ObjectDoesNotExist:
                pass

        report_entity.fields = full_fields

        file = tools.ReportManager.generate_report(report_entity, self.context['language'])

        stream = io.BytesIO()
        file.save(stream)

        file_content = stream.getvalue()

        file_entity = entities.File()
        file_entity.name = type_name + '.xlsx'
        file_entity.size = len(file_content)
        file_entity.content_type = 'xlsx'
        file_entity.content = file_content
        file_entity.thumbnail = False

        file_id = file_gateway.create(file_entity)

        models.FileLink.objects.filter(record_id=report.id,
                                       file_class=self.get_book_record('classes', 'report'),
                                       type=self.get_book_record('file_types', 'report')).delete()

        file_link = models.FileLink()
        file_link.file = models.File.objects.get(pk=file_id)
        file_link.type = self.get_book_record('file_types', 'report')
        file_link.file_class = self.get_book_record('classes', 'report')
        file_link.record_id = report.id
        file_link.save()

        return report.id
